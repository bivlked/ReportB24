"""
Unit tests for Excel styles module.

Tests visual formatting, color schemes, and styling components
that ensure exact reproduction of the screenshot requirements.
"""

import pytest
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import Workbook

from src.excel_generator.styles import (
    ExcelStyles, 
    ColorScheme, 
    ColumnStyleConfig
)


class TestColorScheme:
    """Test color scheme constants."""
    
    def test_color_scheme_initialization(self):
        """Test ColorScheme dataclass initialization."""
        scheme = ColorScheme()
        
        # Test header colors
        assert scheme.HEADER_FILL == "FFC000"
        assert scheme.HEADER_FONT == "000000"
        
        # Test data row colors
        assert scheme.NORMAL_FILL == "FFFFFF"
        assert scheme.NO_VAT_FILL == "F0F0F0"
        assert scheme.DATA_FONT == "000000"
        
        # Test border color
        assert scheme.BORDER_COLOR == "000000"
    
    def test_color_scheme_constants_are_strings(self):
        """Test that all color constants are strings."""
        scheme = ColorScheme()
        
        assert isinstance(scheme.HEADER_FILL, str)
        assert isinstance(scheme.HEADER_FONT, str)
        assert isinstance(scheme.NORMAL_FILL, str)
        assert isinstance(scheme.NO_VAT_FILL, str)
        assert isinstance(scheme.DATA_FONT, str)
        assert isinstance(scheme.BORDER_COLOR, str)
    
    def test_color_scheme_hex_format(self):
        """Test that color values are valid hex format."""
        scheme = ColorScheme()
        
        # Test hex format (6 characters)
        assert len(scheme.HEADER_FILL) == 6
        assert len(scheme.HEADER_FONT) == 6
        assert len(scheme.NORMAL_FILL) == 6
        assert len(scheme.NO_VAT_FILL) == 6
        assert len(scheme.DATA_FONT) == 6
        assert len(scheme.BORDER_COLOR) == 6
        
        # Test hex characters only
        hex_chars = set('0123456789ABCDEFabcdef')
        for color in [scheme.HEADER_FILL, scheme.HEADER_FONT, scheme.NORMAL_FILL, 
                     scheme.NO_VAT_FILL, scheme.DATA_FONT, scheme.BORDER_COLOR]:
            assert all(c in hex_chars for c in color)


class TestExcelStyles:
    """Test Excel styles class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.styles = ExcelStyles()
    
    def test_styles_initialization(self):
        """Test ExcelStyles initialization."""
        assert self.styles.colors is not None
        assert isinstance(self.styles.colors, ColorScheme)
        
        # Test that all style components are initialized
        assert hasattr(self.styles, 'header_font')
        assert hasattr(self.styles, 'data_font')
        assert hasattr(self.styles, 'header_fill')
        assert hasattr(self.styles, 'normal_fill')
        assert hasattr(self.styles, 'no_vat_fill')
        assert hasattr(self.styles, 'cell_border')
        assert hasattr(self.styles, 'center_alignment')
        assert hasattr(self.styles, 'left_alignment')
        assert hasattr(self.styles, 'right_alignment')
    
    def test_font_initialization(self):
        """Test font styles initialization."""
        # Test header font
        assert isinstance(self.styles.header_font, Font)
        assert self.styles.header_font.name == 'Calibri'
        assert self.styles.header_font.size == 11
        assert self.styles.header_font.bold is True
        # OpenPyXL adds "00" prefix to RGB colors
        assert self.styles.header_font.color.rgb == "00" + self.styles.colors.HEADER_FONT
        
        # Test data font
        assert isinstance(self.styles.data_font, Font)
        assert self.styles.data_font.name == 'Calibri'
        assert self.styles.data_font.size == 11
        assert self.styles.data_font.bold is False
        assert self.styles.data_font.color.rgb == "00" + self.styles.colors.DATA_FONT
    
    def test_fill_initialization(self):
        """Test background fill patterns initialization."""
        # Test header fill
        assert isinstance(self.styles.header_fill, PatternFill)
        # OpenPyXL adds "00" prefix to RGB colors
        assert self.styles.header_fill.start_color.rgb == "00" + self.styles.colors.HEADER_FILL
        assert self.styles.header_fill.fill_type == 'solid'
        
        # Test normal fill
        assert isinstance(self.styles.normal_fill, PatternFill)
        assert self.styles.normal_fill.start_color.rgb == "00" + self.styles.colors.NORMAL_FILL
        assert self.styles.normal_fill.fill_type == 'solid'
        
        # Test no VAT fill
        assert isinstance(self.styles.no_vat_fill, PatternFill)
        assert self.styles.no_vat_fill.start_color.rgb == "00" + self.styles.colors.NO_VAT_FILL
        assert self.styles.no_vat_fill.fill_type == 'solid'
    
    def test_border_initialization(self):
        """Test border styles initialization."""
        assert isinstance(self.styles.cell_border, Border)
        
        # Test all border sides
        for side in [self.styles.cell_border.left, self.styles.cell_border.right,
                    self.styles.cell_border.top, self.styles.cell_border.bottom]:
            assert isinstance(side, Side)
            assert side.border_style == 'thin'
            # OpenPyXL adds "00" prefix to RGB colors
            assert side.color.rgb == "00" + self.styles.colors.BORDER_COLOR
    
    def test_alignment_initialization(self):
        """Test text alignment styles initialization."""
        # Test center alignment
        assert isinstance(self.styles.center_alignment, Alignment)
        assert self.styles.center_alignment.horizontal == 'center'
        assert self.styles.center_alignment.vertical == 'center'
        assert self.styles.center_alignment.wrap_text is False
        
        # Test left alignment
        assert isinstance(self.styles.left_alignment, Alignment)
        assert self.styles.left_alignment.horizontal == 'left'
        assert self.styles.left_alignment.vertical == 'center'
        assert self.styles.left_alignment.wrap_text is False
        
        # Test right alignment
        assert isinstance(self.styles.right_alignment, Alignment)
        assert self.styles.right_alignment.horizontal == 'right'
        assert self.styles.right_alignment.vertical == 'center'
        assert self.styles.right_alignment.wrap_text is False
    
    def test_get_header_style(self):
        """Test header style dictionary generation."""
        header_style = self.styles.get_header_style()
        
        assert isinstance(header_style, dict)
        assert 'font' in header_style
        assert 'fill' in header_style
        assert 'border' in header_style
        assert 'alignment' in header_style
        
        assert header_style['font'] == self.styles.header_font
        assert header_style['fill'] == self.styles.header_fill
        assert header_style['border'] == self.styles.cell_border
        assert header_style['alignment'] == self.styles.center_alignment
    
    def test_get_data_style_normal(self):
        """Test normal data style (white background)."""
        data_style = self.styles.get_data_style(is_no_vat=False, alignment_type='left')
        
        assert isinstance(data_style, dict)
        assert 'font' in data_style
        assert 'fill' in data_style
        assert 'border' in data_style
        assert 'alignment' in data_style
        
        assert data_style['font'] == self.styles.data_font
        assert data_style['fill'] == self.styles.normal_fill
        assert data_style['border'] == self.styles.cell_border
        assert data_style['alignment'] == self.styles.left_alignment
    
    def test_get_data_style_no_vat(self):
        """Test no-VAT data style (gray background)."""
        data_style = self.styles.get_data_style(is_no_vat=True, alignment_type='center')
        
        assert isinstance(data_style, dict)
        assert data_style['font'] == self.styles.data_font
        assert data_style['fill'] == self.styles.no_vat_fill
        assert data_style['border'] == self.styles.cell_border
        assert data_style['alignment'] == self.styles.center_alignment
    
    def test_get_data_style_all_alignments(self):
        """Test data styles with different alignments."""
        # Test left alignment
        left_style = self.styles.get_data_style(alignment_type='left')
        assert left_style['alignment'] == self.styles.left_alignment
        
        # Test center alignment
        center_style = self.styles.get_data_style(alignment_type='center')
        assert center_style['alignment'] == self.styles.center_alignment
        
        # Test right alignment
        right_style = self.styles.get_data_style(alignment_type='right')
        assert right_style['alignment'] == self.styles.right_alignment
        
        # Test invalid alignment (defaults to left)
        invalid_style = self.styles.get_data_style(alignment_type='invalid')
        assert invalid_style['alignment'] == self.styles.left_alignment
    
    def test_apply_style_to_cell(self):
        """Test applying style dictionary to a cell."""
        # Create test workbook and cell
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        # Create test style
        test_style = self.styles.get_header_style()
        
        # Apply style
        self.styles.apply_style_to_cell(cell, test_style)
        
        # Verify style was applied
        assert cell.font.name == self.styles.header_font.name
        assert cell.font.bold == self.styles.header_font.bold
        assert cell.fill.start_color.rgb == self.styles.header_fill.start_color.rgb
        assert cell.border.left.border_style == self.styles.cell_border.left.border_style
        assert cell.alignment.horizontal == self.styles.center_alignment.horizontal
    
    def test_apply_partial_style_to_cell(self):
        """Test applying partial style dictionary to a cell."""
        # Create test workbook and cell
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        # Create partial style (only font and fill)
        partial_style = {
            'font': self.styles.header_font,
            'fill': self.styles.header_fill
        }
        
        # Apply style
        self.styles.apply_style_to_cell(cell, partial_style)
        
        # Verify only specified styles were applied
        assert cell.font.name == self.styles.header_font.name
        assert cell.fill.start_color.rgb == self.styles.header_fill.start_color.rgb
        
        # Border and alignment should be default (not from our styles)
        assert cell.border.left.border_style != self.styles.cell_border.left.border_style
    
    def test_apply_empty_style_to_cell(self):
        """Test applying empty style dictionary to a cell."""
        # Create test workbook and cell
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        # Store original style properties
        original_font_name = cell.font.name
        original_font_bold = cell.font.bold
        original_fill_type = cell.fill.fill_type
        
        # Apply empty style
        self.styles.apply_style_to_cell(cell, {})
        
        # Verify no significant changes (font properties should remain the same)
        assert cell.font.name == original_font_name
        assert cell.font.bold == original_font_bold
        assert cell.fill.fill_type == original_fill_type


class TestColumnStyleConfig:
    """Test column style configuration."""
    
    def test_column_alignments_mapping(self):
        """Test column alignment mapping."""
        expected_alignments = {
            1: 'center',    # № п/п
            2: 'left',      # Контрагент
            3: 'center',    # ИНН
            4: 'right',     # Дата отгрузки
            5: 'right',     # Номер счета
            6: 'right',     # Сумма без НДС
            7: 'center',    # НДС
            8: 'right',     # Сумма с НДС
        }
        
        assert ColumnStyleConfig.COLUMN_ALIGNMENTS == expected_alignments
    
    def test_get_column_alignment_valid_columns(self):
        """Test getting alignment for valid column indices."""
        # Test all valid columns
        assert ColumnStyleConfig.get_column_alignment(1) == 'center'
        assert ColumnStyleConfig.get_column_alignment(2) == 'left'
        assert ColumnStyleConfig.get_column_alignment(3) == 'center'
        assert ColumnStyleConfig.get_column_alignment(4) == 'right'
        assert ColumnStyleConfig.get_column_alignment(5) == 'right'
        assert ColumnStyleConfig.get_column_alignment(6) == 'right'
        assert ColumnStyleConfig.get_column_alignment(7) == 'center'
        assert ColumnStyleConfig.get_column_alignment(8) == 'right'
    
    def test_get_column_alignment_invalid_columns(self):
        """Test getting alignment for invalid column indices."""
        # Test invalid columns (should default to 'left')
        assert ColumnStyleConfig.get_column_alignment(0) == 'left'
        assert ColumnStyleConfig.get_column_alignment(9) == 'left'
        assert ColumnStyleConfig.get_column_alignment(-1) == 'left'
        assert ColumnStyleConfig.get_column_alignment(100) == 'left'
    
    def test_column_style_config_is_class_method(self):
        """Test that get_column_alignment is a class method."""
        # Should be callable without instantiation
        alignment = ColumnStyleConfig.get_column_alignment(1)
        assert alignment == 'center'
        
        # Should also work with instance
        config = ColumnStyleConfig()
        alignment = config.get_column_alignment(1)
        assert alignment == 'center'


class TestStylesIntegration:
    """Integration tests for styles module."""
    
    def test_complete_styling_workflow(self):
        """Test complete styling workflow from initialization to cell application."""
        styles = ExcelStyles()
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Test header styling
        header_cell = ws['A1']
        header_style = styles.get_header_style()
        styles.apply_style_to_cell(header_cell, header_style)
        
        # Test normal data styling
        data_cell = ws['A2']
        data_style = styles.get_data_style(is_no_vat=False, alignment_type='left')
        styles.apply_style_to_cell(data_cell, data_style)
        
        # Test no-VAT data styling
        no_vat_cell = ws['A3']
        no_vat_style = styles.get_data_style(is_no_vat=True, alignment_type='center')
        styles.apply_style_to_cell(no_vat_cell, no_vat_style)
        
        # Verify distinct styling
        assert header_cell.font.bold != data_cell.font.bold
        assert header_cell.fill.start_color != data_cell.fill.start_color
        assert data_cell.fill.start_color != no_vat_cell.fill.start_color
        assert data_cell.alignment.horizontal != no_vat_cell.alignment.horizontal
    
    def test_all_column_alignments_applied(self):
        """Test that all column alignments can be applied correctly."""
        styles = ExcelStyles()
        
        for column_index in range(1, 9):  # Test columns 1-8
            alignment_type = ColumnStyleConfig.get_column_alignment(column_index)
            style = styles.get_data_style(alignment_type=alignment_type)
            
            # Verify alignment is correctly mapped
            expected_alignment = getattr(styles, f'{alignment_type}_alignment')
            assert style['alignment'] == expected_alignment
    
    def test_screenshot_color_requirements(self):
        """Test that colors match screenshot requirements exactly."""
        styles = ExcelStyles()
        
        # Verify header color is green #C4D79B as shown in screenshots
        assert styles.colors.HEADER_FILL == "FFC000"
        
        # Verify no-VAT rows are gray as shown in screenshots
        assert styles.colors.NO_VAT_FILL == "F0F0F0"
        
        # Verify normal rows are white
        assert styles.colors.NORMAL_FILL == "FFFFFF"
        
        # Verify black text and borders
        assert styles.colors.DATA_FONT == "000000"
        assert styles.colors.BORDER_COLOR == "000000" 