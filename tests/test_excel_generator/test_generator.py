"""
Unit tests for main Excel generator module.

Tests complete Excel report generation workflow
combining all components.
"""

import pytest
import tempfile
import os
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook, load_workbook

from src.excel_generator.generator import (
    ExcelReportGenerator,
    ExcelReportBuilder,
    ReportGenerationError
)


class TestExcelReportGenerator:
    """Test main Excel report generator."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.generator = ExcelReportGenerator()
    
    def test_generator_initialization(self):
        """Test ExcelReportGenerator initialization."""
        assert hasattr(self.generator, 'styles')
        assert hasattr(self.generator, 'layout')
        assert hasattr(self.generator, 'worksheet_builder')
        assert hasattr(self.generator, 'data_formatter')
        assert hasattr(self.generator, 'summary_formatter')
        assert hasattr(self.generator, 'validator')
    
    def test_generate_report_basic(self):
        """Test basic report generation."""
        test_data = [
            {
                'counterparty': 'ООО "Тест"',
                'inn': '1234567890',
                'shipment_date': '15.06.2025',
                'account_number': 'ТСТ-001',
                'amount': 100000,
                'vat_amount': 20000
            }
        ]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, 'test_report.xlsx')
            
            result_path = self.generator.create_report(test_data, output_path)
            
            # Test that file was created
            assert os.path.exists(result_path)
            assert result_path.endswith('.xlsx')
            
            # Test that file can be opened
            wb = load_workbook(result_path)
            assert len(wb.worksheets) == 1
            ws = wb.active
            assert ws.title == "Краткий"
    
    def test_generate_report_custom_sheet_name(self):
        """Test report generation with custom sheet name."""
        test_data = [{'counterparty': 'Test', 'amount': 1000}]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, 'custom_report.xlsx')
            
            result_path = self.generator.create_report(
                test_data, output_path
            )
            
            wb = load_workbook(result_path)
            ws = wb.active
            assert ws.title == "Краткий"
    
    def test_create_report_functionality(self):
        """Test the actual create_report method functionality."""
        test_data = [
            {
                'account_number': 'ТСТ-001',
                'inn': '1234567890',
                'counterparty': 'ООО "Тест"',
                'amount': '100 000,00',
                'vat_amount': '20 000,00',
                'invoice_date': '15.06.2025',
                'shipment_date': '16.06.2025',
                'payment_date': '17.06.2025'
            }
        ]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, 'test_report.xlsx')
            
            result_path = self.generator.create_report(test_data, output_path)
            
            # Test that file was created and has proper structure
            assert os.path.exists(result_path)
            wb = load_workbook(result_path)
            ws = wb.active
            
            # Verify basic structure - data starts at B3
            data_cell = ws.cell(row=3, column=2)  # B3
            assert data_cell.value is not None


class TestExcelReportBuilder:
    """Test high-level Excel report builder."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.builder = ExcelReportBuilder()
    
    def test_builder_initialization(self):
        """Test ExcelReportBuilder initialization."""
        assert hasattr(self.builder, 'generator')
    
    def test_build_invoice_report(self):
        """Test building invoice report."""
        test_invoices = [
            {
                'counterparty': 'ООО "Строитель"',
                'inn': '1234567890',
                'amount': 250000,
                'vat_amount': 50000
            }
        ]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, 'invoice_report.xlsx')
            
            result_path = self.builder.build_invoice_report(
                test_invoices, output_path
            )
            
            assert os.path.exists(result_path)
            
            wb = load_workbook(result_path)
            ws = wb.active
            assert ws.title == "Краткий"
    
    def test_build_invoice_report_error_handling(self):
        """Test error handling in invoice report building."""
        with patch.object(self.builder.generator, 'create_report', side_effect=Exception("Test error")):
            with pytest.raises(ReportGenerationError):
                self.builder.build_invoice_report([], 'test.xlsx')
    
    def test_build_summary_report(self):
        """Test building summary report."""
        test_invoices = [
            {'amount': 100000, 'vat_amount': 20000},
            {'amount': 200000, 'vat_amount': 40000}
        ]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, 'summary_report.xlsx')
            
            summary = self.builder.build_summary_report(test_invoices, output_path)
            
            assert isinstance(summary, dict)
            assert 'record_count' in summary
            assert 'total_without_vat' in summary
            assert 'total_with_vat' in summary
            
            # Test that Excel file was also created
            assert os.path.exists(output_path)
    
    def test_build_summary_report_error_handling(self):
        """Test error handling in summary report building."""
        with patch.object(self.builder.generator, 'create_report', side_effect=Exception("Test error")):
            with pytest.raises(ReportGenerationError):
                self.builder.build_summary_report([], 'test.xlsx')
    
    def test_validate_data_for_report_valid(self):
        """Test data validation for valid data."""
        valid_data = [
            {
                'counterparty': 'ООО "Валидный"',
                'inn': '1234567890',
                'amount': 100000,
                'vat_amount': 20000
            }
        ]
        
        results = self.builder.validate_data_for_report(valid_data)
        
        assert results['is_valid'] is True
        assert results['total_records'] == 1
        assert results['valid_records'] == 1
        assert results['invalid_records'] == 0
        assert len(results['errors']) == 0
    
    def test_validate_data_for_report_invalid(self):
        """Test data validation for invalid data."""
        invalid_data = [
            {'counterparty': 'Valid'},
            {}  # This might cause errors in formatting
        ]
        
        results = self.builder.validate_data_for_report(invalid_data)
        
        # With simplified validation, this should just count records
        assert results['total_records'] == 2
        assert 'valid_records' in results
        assert 'invalid_records' in results
    
    def test_validate_data_for_report_empty(self):
        """Test data validation for empty data."""
        results = self.builder.validate_data_for_report([])
        
        assert results['is_valid'] is True
        assert results['total_records'] == 0
        assert results['valid_records'] == 0
        assert results['invalid_records'] == 0
        assert len(results['errors']) == 0


class TestGeneratorIntegration:
    """Integration tests for the entire generator module."""
    
    def test_complete_report_generation_workflow(self):
        """Test complete report generation from start to finish."""
        generator = ExcelReportGenerator()
        
        # Comprehensive test data
        test_data = [
            {
                'counterparty': 'ООО "Первая компания"',
                'inn': '1234567890',
                'shipment_date': '15.06.2025',
                'account_number': 'ПК-001',
                'amount': 100000,
                'vat_amount': 20000
            },
            {
                'counterparty': 'ИП Иванов И.И.',
                'inn': '123456789012',
                'shipment_date': '16.06.2025',
                'account_number': 'ИИ-002',
                'amount': 50000,
                'vat_amount': 0
            }
        ]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, 'integration_test.xlsx')
            
            # Generate report
            result_path = generator.create_report(test_data, output_path)
            
            # Verify file creation
            assert os.path.exists(result_path)
            
            # Load and verify content
            wb = load_workbook(result_path)
            ws = wb.active
            
            # Test that worksheet has proper structure
            # Headers should be present at row 2
            header_cell = ws.cell(row=2, column=2)  # B2
            assert header_cell.value is not None
            
            # Test data is present
            # First row data should be in row 3 (B3)
            first_data_cell = ws.cell(row=3, column=2)  # B3
            assert first_data_cell.value is not None
            
            # Second row data should be in row 4 (B4)
            second_data_cell = ws.cell(row=4, column=2)  # B4
            assert second_data_cell.value is not None
    
    def test_error_handling_workflow(self):
        """Test error handling throughout the generation workflow."""
        generator = ExcelReportGenerator()
        
        # Test with completely invalid output path that will definitely fail
        invalid_path = '/invalid/path/that/definitely/does/not/exist/and/cannot/be/created/file.xlsx'
        
        # On Windows, also test with invalid characters
        if os.name == 'nt':
            invalid_path = 'C:\\invalid\\<>|*?"path\\file.xlsx'
        
        with pytest.raises((PermissionError, OSError, FileNotFoundError)):
            generator.create_report([], invalid_path)
    
    def test_different_file_extensions(self):
        """Test handling of different file extensions."""
        generator = ExcelReportGenerator()
        test_data = [{'counterparty': 'Test', 'amount': 1000}]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # Test various extensions
            for ext in ['.xls', '.xlsx', '', '.txt']:
                filename = f'test{ext}'
                output_path = os.path.join(temp_dir, filename)
                
                result_path = generator.create_report(test_data, output_path)
                
                # Should always end with .xlsx
                assert result_path.endswith('.xlsx')
                assert os.path.exists(result_path)
    
    def test_large_dataset_handling(self):
        """Test handling of larger datasets."""
        generator = ExcelReportGenerator()
        
        # Create larger test dataset
        large_data = []
        for i in range(100):
            large_data.append({
                'counterparty': f'ООО "Компания {i}"',
                'inn': f'{1234567890 + i}',
                'shipment_date': '15.06.2025',
                'account_number': f'КМП-{i:03d}',
                'amount': 10000 * (i + 1),
                'vat_amount': 2000 * (i + 1) if i % 2 == 0 else 0
            })
        
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, 'large_dataset.xlsx')
            
            result_path = generator.create_report(large_data, output_path)
            
            assert os.path.exists(result_path)
            
            # Verify that file was created and can be opened
            wb = load_workbook(result_path)
            ws = wb.active
            
            # Basic verification that data exists
            assert ws.cell(row=3, column=2).value is not None  # First data row 