import requests
from datetime import datetime
import time
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Вебхук и метод
webhook_url = "https://softway.bitrix24.ru/rest/12/kt17hzrauafyfem0/"
method = "crm.item.list"

def get_company_info_by_invoice(invoice_number):
    """
    Ищем счёт по номеру (accountNumber), смотрим привязку в crm.requisite.link,
    достаём реквизит (RQ_INN, RQ_COMPANY_NAME, RQ_NAME).
    - Если ИНН=10 -> (RQ_COMPANY_NAME, RQ_INN).
    - Если ИНН=12 -> ("ИП <RQ_NAME>", RQ_INN).
    - Иначе (RQ_COMPANY_NAME, RQ_INN).
    """
    try:
        resp = requests.get(f"{webhook_url}crm.item.list", params={
            'filter[accountNumber]': invoice_number,
            'entityTypeId': 31
        })
        time.sleep(0.5)
        data = resp.json()

        items = data.get('result', {}).get('items', [])
        if not items:
            raise Exception(f"Счёт №{invoice_number} не найден. Ответ: {data}")

        inv_item = items[0]
        inv_id = inv_item.get('id')
        if not inv_id:
            raise Exception(f"Нет поля 'id' у счёта №{invoice_number}.")

        # Привязка crm.requisite.link.list
        link_resp = requests.get(f"{webhook_url}crm.requisite.link.list", params={
            'filter[ENTITY_TYPE_ID]': 31,
            'filter[ENTITY_ID]': inv_id
        })
        time.sleep(0.5)
        link_data = link_resp.json()
        link_items = link_data.get('result', [])
        if not link_items:
            raise Exception(f"Нет реквизитов (crm.requisite.link) у счёта ID={inv_id}.")

        req_id = link_items[0].get('REQUISITE_ID')
        if not req_id or int(req_id) <= 0:
            raise Exception(f"Некорректный REQUISITE_ID={req_id}.")

        # Достаём реквизит
        req_resp = requests.post(f"{webhook_url}crm.requisite.get", json={"id": str(req_id)})
        time.sleep(0.5)
        req_data = req_resp.json()
        if not req_data.get('result'):
            raise Exception(f"Не удалось получить реквизит ID={req_id}. {req_data}")

        fields = req_data['result']
        rq_inn = fields.get('RQ_INN', '')
        rq_company = fields.get('RQ_COMPANY_NAME', '')
        rq_name = fields.get('RQ_NAME', '')

        if rq_inn.isdigit():
            if len(rq_inn) == 10:
                return rq_company, rq_inn
            elif len(rq_inn) == 12:
                return (f"ИП {rq_name}" if rq_name else "ИП (нет имени)", rq_inn)
            else:
                return rq_company, rq_inn
        else:
            return rq_company, rq_inn

    except Exception as e:
        print("Ошибка:", e)
        return None, None

def get_all_smart_invoices(start_date, end_date):
    """
    Получаем список Smart Invoices,
    фильтруем по дате отгрузки UFCRM_SMART_INVOICE_1651168135187 (между start_date и end_date).
    """
    url = f"{webhook_url}{method}"
    all_invoices = []
    start = 0

    while True:
        params = {
            'entityTypeId': 31,
            'start': start,
            'filter': {
                '!stageId': 'DT31_1:D'
            },
            'select': [
                'id',
                'accountNumber',
                'statusId',
                'dateBill',
                'price',
                'UFCRM_SMART_INVOICE_1651168135187',  # Дата отгрузки
                'UFCRM_626D6ABE98692',               # Дата оплаты
                'begindate',
                'opportunity',
                'stageId',
                'taxValue'
            ]
        }
        resp = requests.post(url, json=params)
        time.sleep(0.5)
        if resp.status_code == 200:
            data = resp.json()
            if 'result' in data and 'items' in data['result']:
                all_invoices.extend(data['result']['items'])
                if 'next' in data:
                    start = data['next']
                else:
                    break
            else:
                break
        else:
            print("Ошибка при запросе:", resp.status_code, resp.text)
            break

    filtered = []
    for inv in all_invoices:
        ship_date_str = inv.get('UFCRM_SMART_INVOICE_1651168135187')
        if ship_date_str:
            try:
                d = datetime.fromisoformat(ship_date_str.replace('Z', '+00:00')).date()
                if start_date <= d <= end_date:
                    filtered.append(inv)
            except ValueError as ex:
                print(f"Ошибка преобразования даты отгрузки (ID={inv['id']}):", ex)

    return filtered

def format_date(date_str):
    if not date_str:
        return ""
    d = datetime.fromisoformat(date_str.replace('Z', '+00:00')).date()
    return d.strftime("%d.%m.%Y")

def format_currency(amount):
    return f"{amount:,.2f}".replace(',', ' ').replace('.', ',')

def main():
    # Ввод дат (DD.MM.YYYY)
    start_date_str = input("Введите дату начала периода отгрузки (DD.MM.YYYY): ")
    end_date_str = input("Введите дату конца периода отгрузки (DD.MM.YYYY): ")
    file_name = input("Введите желаемое имя файла (без расширения): ") + ".xlsx"
    start_date = datetime.strptime(start_date_str, "%d.%m.%Y").date()
    end_date = datetime.strptime(end_date_str, "%d.%m.%Y").date()

    # Получаем счета
    invoices = get_all_smart_invoices(start_date, end_date)

    # Шаг 1: Создаём Excel-файл и пишем строки
    wb = Workbook()
    ws = wb.active
    ws.title = "Счета (Smart Invoices)"

    headers = ["Номер", "ИНН", "Контрагент", "Сумма", "НДС", "Дата счёта", "Дата отгрузки", "Дата оплаты"]
    ws.append(headers)

    # Заливки для строк
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    grey_fill  = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    row_index = 2  # т.к. первая строка занята заголовками
    for inv in invoices:
        acc_num = inv.get('accountNumber', '')
        sum_val = format_currency(float(inv.get('opportunity', 0)))
        tax_val = float(inv.get('taxValue', 0))
        tax_text = format_currency(tax_val) if tax_val != 0 else "нет"

        date_bill_str = inv.get('begindate')
        date_bill = format_date(date_bill_str)

        ship_date_str = inv.get('UFCRM_SMART_INVOICE_1651168135187')
        ship_date = format_date(ship_date_str)

        pay_date_str = inv.get('UFCRM_626D6ABE98692')
        pay_date = format_date(pay_date_str) if pay_date_str else ""

        comp_name, inn = get_company_info_by_invoice(acc_num)
        if not comp_name and not inn:
            comp_name, inn = "Не найдено", "Не найдено"

        row_data = [acc_num, inn, comp_name, sum_val, tax_text, date_bill, ship_date, pay_date]
        ws.append(row_data)

        # Логика заливки:
        # если нет даты оплаты => красная строка
        # иначе, если tax_text="нет" => серая строка
        # иначе белая
        if pay_date == "":
            fill_color = red_fill
        elif tax_text == "нет":
            fill_color = grey_fill
        else:
            fill_color = white_fill

        for col_idx in range(1, len(headers)+1):
            cell = ws.cell(row=row_index, column=col_idx)
            cell.fill = fill_color

        row_index += 1

    # Временный файл
    temp_file = "Отчёт_Счета.xlsx"
    wb.save(temp_file)

    # Шаг 2: Открываем заново для форматирования (ширина столбцов, выравнивание, границы)
    wb2 = load_workbook(temp_file)
    ws2 = wb2.active

    # Автоподбор ширины столбцов
    for col in ws2.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                length = len(str(val))
                if length > max_length:
                    max_length = length
        ws2.column_dimensions[col_letter].width = max_length + 2

    # Выравнивания
    center_alignment = Alignment(horizontal="center")
    left_alignment   = Alignment(horizontal="left")
    right_alignment  = Alignment(horizontal="right")

    # Заголовок (первая строка) выравниваем по центру и задаём цвет
    header_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")

    for cell in ws2["1:1"]:
        cell.alignment = center_alignment
        cell.fill = header_fill

    # Пример, как центрировать целые столбцы
    # B, C, E — center
    for cell in ws2["B:B"]:
        cell.alignment = center_alignment
    for cell in ws2["C:C"]:
        cell.alignment = center_alignment
    for cell in ws2["E:E"]:
        cell.alignment = center_alignment

    # F, G, H — выравнивание вправо
    for cell in ws2["F:F"]:
        cell.alignment = right_alignment
    for cell in ws2["G:G"]:
        cell.alignment = right_alignment
    for cell in ws2["H:H"]:
        cell.alignment = right_alignment

    # Для D:D (Контрагент) — сдвиг влево, но только со второй строки
    for cell in ws2["D:D"]:
        if cell.row > 1:
            cell.alignment = left_alignment

    # Добавим тонкие границы для всех ячеек
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws2.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Сохраняем окончательно
    wb2.save(file_name)

    print(f"Данные успешно сохранены в файл {file_name} с форматированием.")


if __name__ == "__main__":
    main()
