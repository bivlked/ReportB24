"""
Excel styles module for Bitrix24 report generation.

Implements visual formatting that exactly matches the provided screenshots:
- Orange header background #FFC000 (as shown in screenshot 01.png)
- Gray rows for "Без НДС" items  
- White background for normal rows
- Proper alignment and borders
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import Dict, Any
from dataclasses import dataclass


@dataclass
class ColorScheme:
    """Color scheme constants matching the screenshot requirements."""
    
    # Header colors (orange background as shown in screenshots)
    HEADER_FILL = "FFC000"  # Orange for headers (matching screenshot)
    HEADER_FONT = "000000"  # Black text on orange background
    
    # Data row colors
    NORMAL_FILL = "FFFFFF"  # White background for normal rows
    NO_VAT_FILL = "D3D3D3"  # Gray for "Без НДС" rows (matching ShortReport.py)
    UNPAID_FILL = "FFC0CB"  # Light red/pink for unpaid invoices
    DATA_FONT = "000000"    # Black text for data
    
    # Border color
    BORDER_COLOR = "000000"  # Black borders


class ExcelStyles:
    """
    Excel formatting styles for the Bitrix24 report.
    
    Provides consistent styling throughout the report that matches
    the visual requirements from the provided screenshots.
    """
    
    def __init__(self):
        self.colors = ColorScheme()
        self._init_fonts()
        self._init_fills()
        self._init_borders()
        self._init_alignments()
    
    def _init_fonts(self) -> None:
        """Initialize font styles."""
        # Header font: bold, size 11 (standard Excel)
        self.header_font = Font(
            name='Calibri',
            size=11,
            bold=True,
            color=self.colors.HEADER_FONT
        )
        
        # Data font: normal, size 11
        self.data_font = Font(
            name='Calibri',
            size=11,
            bold=False,
            color=self.colors.DATA_FONT
        )
    
    def _init_fills(self) -> None:
        """Initialize background fill patterns."""
        # Header fill: orange background (matching screenshots)
        self.header_fill = PatternFill(
            start_color=self.colors.HEADER_FILL,
            end_color=self.colors.HEADER_FILL,
            fill_type='solid'
        )
        
        # Normal row fill: white background
        self.normal_fill = PatternFill(
            start_color=self.colors.NORMAL_FILL,
            end_color=self.colors.NORMAL_FILL,
            fill_type='solid'
        )
        
        # No VAT row fill: gray background
        self.no_vat_fill = PatternFill(
            start_color=self.colors.NO_VAT_FILL,
            end_color=self.colors.NO_VAT_FILL,
            fill_type='solid'
        )
        
        # Unpaid row fill: light red background
        self.unpaid_fill = PatternFill(
            start_color=self.colors.UNPAID_FILL,
            end_color=self.colors.UNPAID_FILL,
            fill_type='solid'
        )
    
    def _init_borders(self) -> None:
        """Initialize border styles."""
        # Thin black border for all cells
        thin_side = Side(border_style="thin", color=self.colors.BORDER_COLOR)
        self.cell_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )
        
        # Thick bottom border for invoice separators
        thick_side = Side(border_style="thick", color=self.colors.BORDER_COLOR)
        self.invoice_separator_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thick_side  # Толстая нижняя граница
        )
    
    def _init_alignments(self) -> None:
        """Initialize text alignment styles."""
        # Center alignment for headers and INN/VAT columns
        self.center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=False
        )
        
        # Left alignment for contractor names
        self.left_alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=False
        )
        
        # Right alignment for dates and amounts
        self.right_alignment = Alignment(
            horizontal='right',
            vertical='center',
            wrap_text=False
        )
    
    def get_header_style(self) -> Dict[str, Any]:
        """
        Get complete header cell style.
        
        Returns:
            Dict containing all header style properties
        """
        return {
            'font': self.header_font,
            'fill': self.header_fill,
            'border': self.cell_border,
            'alignment': self.center_alignment
        }
    
    def get_data_style(self, is_no_vat: bool = False, is_unpaid: bool = False,
                      alignment_type: str = 'left') -> Dict[str, Any]:
        """
        Get complete data cell style.
        
        Args:
            is_no_vat: True if this is a "Без НДС" row (gray background)
            is_unpaid: True if this is an unpaid invoice (red background)
            alignment_type: 'left', 'center', or 'right'
        
        Returns:
            Dict containing all data cell style properties
        """
        # Select fill based on payment and VAT status
        # Priority: unpaid > no_vat > normal
        if is_unpaid:
            fill = self.unpaid_fill
        elif is_no_vat:
            fill = self.no_vat_fill
        else:
            fill = self.normal_fill
        
        # Select alignment
        alignment_map = {
            'left': self.left_alignment,
            'center': self.center_alignment,
            'right': self.right_alignment
        }
        alignment = alignment_map.get(alignment_type, self.left_alignment)
        
        return {
            'font': self.data_font,
            'fill': fill,
            'border': self.cell_border,
            'alignment': alignment
        }
    
    def apply_style_to_cell(self, cell, style_dict: Dict[str, Any]) -> None:
        """
        Apply style dictionary to a cell.
        
        Args:
            cell: OpenPyXL cell object
            style_dict: Style properties dictionary
        """
        if 'font' in style_dict:
            cell.font = style_dict['font']
        if 'fill' in style_dict:
            cell.fill = style_dict['fill']
        if 'border' in style_dict:
            cell.border = style_dict['border']
        if 'alignment' in style_dict:
            cell.alignment = style_dict['alignment']


class ColumnStyleConfig:
    """
    Configuration for column-specific styling.
    
    Defines which alignment and formatting each column should use
    based on the screenshot requirements.
    """
    
    # Column alignment mapping (1-based column numbers)
    # Updated for new column headers: Номер, ИНН, Контрагент, Сумма, НДС, Дата счёта, Дата отгрузки, Дата оплаты
    COLUMN_ALIGNMENTS = {
        1: 'center',    # Номер (center)
        2: 'center',    # ИНН (center)
        3: 'left',      # Контрагент (left)
        4: 'right',     # Сумма (right)
        5: 'center',    # НДС (center)
        6: 'right',     # Дата счёта (right)
        7: 'right',     # Дата отгрузки (right)
        8: 'right',     # Дата оплаты (right)
    }
    
    @classmethod
    def get_column_alignment(cls, column_index: int) -> str:
        """
        Get alignment for specific column.
        
        Args:
            column_index: 1-based column index
            
        Returns:
            Alignment type: 'left', 'center', or 'right'
        """
        return cls.COLUMN_ALIGNMENTS.get(column_index, 'left') 