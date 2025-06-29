#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Правильный генератор отчетов Bitrix24 Excel
Основан на проверенном ShortReport.py с использованием config.ini

Использует правильные API методы:
- crm.item.list для Smart Invoices (entityTypeId=31)
- crm.requisite.link.list для связей реквизитов
- crm.requisite.get для получения ИНН и названий компаний
"""

import requests
import configparser
import time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class EnhancedBitrix24Reporter:
    """Правильный генератор отчетов Bitrix24 с полными реквизитами"""
    
    def __init__(self, config_path="config.ini"):
        """Инициализация с загрузкой конфигурации"""
        print(f"⏰ Время инициализации: {datetime.now().strftime('%Y.%m.%d %H:%M:%S')}")
        self.config = configparser.ConfigParser()
        self.config.read(config_path, encoding='utf-8')
        
        # Получаем настройки из config.ini (правильные секции)
        self.webhook_url = self.config.get('BitrixAPI', 'webhookurl')
        self.save_folder = self.config.get('AppSettings', 'defaultsavefolder', fallback='reports')
        self.file_name = self.config.get('AppSettings', 'defaultfilename', fallback='Enhanced_Report.xlsx')
        
        # Период из конфига (с правильными датами для 2022)
        start_date_str = self.config.get('ReportPeriod', 'startdate', fallback='01.01.2022')
        end_date_str = self.config.get('ReportPeriod', 'enddate', fallback='31.12.2022')
        
        self.start_date = datetime.strptime(start_date_str, "%d.%m.%Y").date()
        self.end_date = datetime.strptime(end_date_str, "%d.%m.%Y").date()
        
        # Создаем папку для отчетов
        Path(self.save_folder).mkdir(exist_ok=True)
        
        print(f"🔗 Webhook URL: {self.webhook_url}")
        print(f"📅 Период: {start_date_str} - {end_date_str}")
        print(f"📁 Папка сохранения: {self.save_folder}")
        print(f"📄 Имя файла: {self.file_name}")

    def get_company_info_by_invoice(self, invoice_number):
        """
        Получение реквизитов контрагента по номеру счета
        Точная копия логики из ShortReport.py
        """
        try:
            # 1. Ищем счёт по номеру (accountNumber)
            resp = requests.get(f"{self.webhook_url}crm.item.list", params={
                'filter[accountNumber]': invoice_number,
                'entityTypeId': 31
            })
            time.sleep(0.5)  # Rate limiting как в ShortReport.py
            data = resp.json()

            items = data.get('result', {}).get('items', [])
            if not items:
                print(f"⚠️ Счёт №{invoice_number} не найден")
                return None, None

            inv_item = items[0]
            inv_id = inv_item.get('id')
            if not inv_id:
                print(f"⚠️ Нет поля 'id' у счёта №{invoice_number}")
                return None, None

            # 2. Ищем привязку в crm.requisite.link
            link_resp = requests.get(f"{self.webhook_url}crm.requisite.link.list", params={
                'filter[ENTITY_TYPE_ID]': 31,
                'filter[ENTITY_ID]': inv_id
            })
            time.sleep(0.5)
            link_data = link_resp.json()
            link_items = link_data.get('result', [])
            if not link_items:
                print(f"⚠️ Нет реквизитов у счёта ID={inv_id}")
                return "Нет реквизитов", "Нет реквизитов"

            req_id = link_items[0].get('REQUISITE_ID')
            if not req_id or int(req_id) <= 0:
                print(f"⚠️ Некорректный REQUISITE_ID={req_id}")
                return "Некорректный реквизит", "Некорректный реквизит"

            # 3. Получаем реквизит
            req_resp = requests.post(f"{self.webhook_url}crm.requisite.get", json={"id": str(req_id)})
            time.sleep(0.5)
            req_data = req_resp.json()
            if not req_data.get('result'):
                print(f"⚠️ Не удалось получить реквизит ID={req_id}")
                return "Ошибка реквизита", "Ошибка реквизита"

            fields = req_data['result']
            rq_inn = fields.get('RQ_INN', '')
            rq_company = fields.get('RQ_COMPANY_NAME', '')
            rq_name = fields.get('RQ_NAME', '')

            # 4. Логика определения типа по ИНН (как в ShortReport.py)
            if rq_inn.isdigit():
                if len(rq_inn) == 10:
                    return rq_company, rq_inn  # ООО/ЗАО
                elif len(rq_inn) == 12:
                    return (f"ИП {rq_name}" if rq_name else "ИП (нет имени)", rq_inn)  # ИП
                else:
                    return rq_company, rq_inn
            else:
                return rq_company, rq_inn

        except Exception as e:
            print(f"❌ Ошибка получения реквизитов для счёта {invoice_number}: {e}")
            return "Ошибка", "Ошибка"

    def get_all_smart_invoices(self):
        """
        Получение всех Smart Invoices с фильтрацией по дате отгрузки
        Точная копия логики из ShortReport.py
        """
        print(f"🔍 Получение Smart Invoices за период {self.start_date} - {self.end_date}")
        url = f"{self.webhook_url}crm.item.list"
        all_invoices = []
        start = 0
        page = 1

        while True:
            print(f"📄 Обработка страницы {page}...")
            params = {
                'entityTypeId': 31,
                'start': start,
                'filter': {
                    '!stageId': 'DT31_1:D'  # исключаем удаленные
                },
                'select': [
                    'id',
                    'accountNumber',
                    'statusId',
                    'dateBill',
                    'price',
                    'UFCRM_SMART_INVOICE_1651168135187',  # Дата отгрузки
                    'UFCRM_626D6ABE98692',               # Дата оплаты
                    'begindate',
                    'opportunity',
                    'stageId',
                    'taxValue'
                ]
            }
            
            resp = requests.post(url, json=params)
            time.sleep(0.5)  # Rate limiting
            
            if resp.status_code == 200:
                data = resp.json()
                if 'result' in data and 'items' in data['result']:
                    page_items = data['result']['items']
                    all_invoices.extend(page_items)
                    print(f"✅ Получено {len(page_items)} записей на странице {page}")
                    
                    if 'next' in data:
                        start = data['next']
                        page += 1
                    else:
                        break
                else:
                    print("⚠️ Нет данных в ответе API")
                    break
            else:
                print(f"❌ Ошибка API: {resp.status_code} - {resp.text}")
                break

        print(f"📊 Всего получено записей: {len(all_invoices)}")

        # Фильтрация по дате отгрузки (как в ShortReport.py)
        filtered = []
        for inv in all_invoices:
            ship_date_str = inv.get('UFCRM_SMART_INVOICE_1651168135187')
            if ship_date_str:
                try:
                    d = datetime.fromisoformat(ship_date_str.replace('Z', '+00:00')).date()
                    if self.start_date <= d <= self.end_date:
                        filtered.append(inv)
                except ValueError as ex:
                    print(f"⚠️ Ошибка даты отгрузки (ID={inv['id']}): {ex}")

        print(f"🎯 Отфильтровано по дате отгрузки: {len(filtered)} записей")
        return filtered

    def format_date(self, date_str):
        """Форматирование даты (как в ShortReport.py)"""
        if not date_str:
            return ""
        try:
            d = datetime.fromisoformat(date_str.replace('Z', '+00:00')).date()
            return d.strftime("%d.%m.%Y")
        except:
            return ""

    def format_currency(self, amount):
        """Форматирование валюты (как в ShortReport.py)"""
        try:
            return f"{float(amount):,.2f}".replace(',', ' ').replace('.', ',')
        except:
            return "0,00"

    def generate_report(self):
        """Генерация полного отчета с правильным форматированием"""
        print("🚀 Запуск генерации отчета...")
        
        # Получаем данные
        invoices = self.get_all_smart_invoices()
        if not invoices:
            print("❌ Нет данных для отчета")
            return False

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Счета (Smart Invoices)"

        # Правильные заголовки как в ShortReport.py
        headers = ["Номер", "ИНН", "Контрагент", "Сумма", "НДС", "Дата счёта", "Дата отгрузки", "Дата оплаты"]
        ws.append(headers)

        # Настройка цветов (как в ShortReport.py)
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        header_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")

        print(f"📝 Обработка {len(invoices)} записей...")
        row_index = 2
        
        for i, inv in enumerate(invoices, 1):
            if i % 10 == 0:
                print(f"⏳ Обработано {i}/{len(invoices)} записей...")
                
            acc_num = inv.get('accountNumber', '')
            sum_val = self.format_currency(inv.get('opportunity', 0))
            tax_val = float(inv.get('taxValue', 0))
            tax_text = self.format_currency(tax_val) if tax_val != 0 else "нет"

            date_bill = self.format_date(inv.get('begindate'))
            ship_date = self.format_date(inv.get('UFCRM_SMART_INVOICE_1651168135187'))
            pay_date_str = inv.get('UFCRM_626D6ABE98692')
            pay_date = self.format_date(pay_date_str) if pay_date_str else ""

            # Получаем реквизиты (самая важная часть!)
            comp_name, inn = self.get_company_info_by_invoice(acc_num)
            if not comp_name and not inn:
                comp_name, inn = "Не найдено", "Не найдено"

            row_data = [acc_num, inn, comp_name, sum_val, tax_text, date_bill, ship_date, pay_date]
            ws.append(row_data)

            # Цветовая логика (как в ShortReport.py)
            if pay_date == "":
                fill_color = red_fill  # неоплаченные
            elif tax_text == "нет":
                fill_color = grey_fill  # без НДС
            else:
                fill_color = white_fill  # обычные

            # Применяем цвет ко всей строке
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_index, column=col_idx)
                cell.fill = fill_color

            row_index += 1

        # Сохраняем временный файл
        temp_path = Path(self.save_folder) / "temp_report.xlsx"
        wb.save(temp_path)

        # Второй этап: профессиональное форматирование (как в ShortReport.py)
        print("🎨 Применение профессионального форматирования...")
        wb2 = load_workbook(temp_path)
        ws2 = wb2.active

        # Автоподбор ширины столбцов
        for col in ws2.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val is not None:
                    length = len(str(val))
                    if length > max_length:
                        max_length = length
            ws2.column_dimensions[col_letter].width = max_length + 2

        # Выравнивания
        center_alignment = Alignment(horizontal="center")
        left_alignment = Alignment(horizontal="left")
        right_alignment = Alignment(horizontal="right")

        # Заголовок (первая строка) - зеленый и по центру
        for cell in ws2["1:1"]:
            cell.alignment = center_alignment
            cell.fill = header_fill

        # Выравнивание столбцов (как в ShortReport.py)
        for cell in ws2["B:B"]:  # ИНН - центр
            cell.alignment = center_alignment
        for cell in ws2["E:E"]:  # НДС - центр
            cell.alignment = center_alignment
        for cell in ws2["F:F"]:  # Дата счёта - вправо
            cell.alignment = right_alignment
        for cell in ws2["G:G"]:  # Дата отгрузки - вправо
            cell.alignment = right_alignment
        for cell in ws2["H:H"]:  # Дата оплаты - вправо
            cell.alignment = right_alignment
        for cell in ws2["D:D"]:  # Контрагент - влево (только данные)
            if cell.row > 1:
                cell.alignment = left_alignment

        # Границы для всех ячеек
        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for row in ws2.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Сохраняем итоговый файл
        final_path = Path(self.save_folder) / self.file_name
        wb2.save(final_path)
        
        # Удаляем временный файл
        temp_path.unlink()
        
        print(f"✅ Отчет успешно создан: {final_path}")
        print(f"📊 Обработано записей: {len(invoices)}")
        return True


def main():
    """Главная функция запуска"""
    print("=" * 50)
    print("🎯 ГЕНЕРАТОР ОТЧЕТОВ BITRIX24 (ENHANCED)")
    print("=" * 50)
    
    try:
        reporter = EnhancedBitrix24Reporter()
        success = reporter.generate_report()
        
        if success:
            print("\n🎉 ОТЧЕТ УСПЕШНО СОЗДАН!")
        else:
            print("\n❌ ОШИБКА СОЗДАНИЯ ОТЧЕТА")
            
    except Exception as e:
        print(f"\n💥 Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
    
    input("\n⏸️ Нажмите Enter для выхода...")


if __name__ == "__main__":
    main() 