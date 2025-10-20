#!/usr/bin/env python3
"""
Генератор Excel отчетов для Bitrix24.

Создает финальные отчеты с правильным форматированием,
соответствующим скриншотам-эталонам.
"""

from typing import List, Dict, Any, Optional
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging

from .styles import ExcelStyles, ColumnStyleConfig
from .layout import WorksheetBuilder, ReportLayout, MultiSheetBuilder, DetailedWorksheetBuilder
from .formatter import ExcelDataFormatter, ExcelSummaryFormatter, DataValidator


class ExcelReportGenerator:
    """
    Финальный генератор Excel отчетов.
    
    Создает отчеты с полным соответствием скриншотам-эталонам:
    - Отступы сверху и слева
    - Заморозка заголовков при прокручивании 
    - Правильные цвета и границы
    - Выравнивание по типу данных
    - Числовое форматирование
    - Итоги как на скриншоте 04.png
    """
    
    def __init__(self):
        self.styles = ExcelStyles()
        self.layout = ReportLayout()
        self.worksheet_builder = WorksheetBuilder()
        self.data_formatter = ExcelDataFormatter()
        self.summary_formatter = ExcelSummaryFormatter()
        self.validator = DataValidator()
        self.logger = logging.getLogger(__name__)
        
        # Настройки отступов (начинаем с B2 вместо A1)
        self.start_row = 2  # Отступ сверху
        self.start_col = 2  # Отступ слева (столбец B)
        
        # Новые компоненты для детального отчета - Фаза 4
        self.multi_sheet_builder = MultiSheetBuilder()
        self.detailed_builder = DetailedWorksheetBuilder()
    
    def create_report(self, data: List[Dict[str, Any]], output_path: str) -> str:
        """
        Создает финальный Excel отчет с полным форматированием.
        
        Args:
            data: Обработанные данные счетов
            output_path: Путь для сохранения файла
            
        Returns:
            Путь к созданному файлу
        """
        try:
            # Обеспечиваем правильное расширение .xlsx
            output_path = self._ensure_xlsx_extension(output_path)
            
            self.logger.info(f"📊 Создание финального Excel отчета: {len(data)} записей")
            
            # Создаем новую книгу
            wb = Workbook()  
            ws = wb.active
            ws.title = "Краткий"  # 7. Название листа "Краткий"
            
            # 1. Добавляем заголовки с форматированием
            self._add_headers(ws)
            
            # 2. Добавляем данные с правильным форматированием
            self._add_data_rows(ws, data)
            
            # 3. Применяем границы только вокруг данных (без итогов)
            self._apply_data_table_borders(ws, len(data))
            
            # 4. Добавляем итоги как на скриншоте 04.png (ВНЕ жирной рамки)
            self._add_summary_section_new_format(ws, data)
            
            # 5. Заморозка заголовков
            self._freeze_headers(ws)
            
            # 6. Настройка столбцов с автошириной
            self._adjust_column_widths_auto(ws, data)
            
            # Сохраняем файл
            wb.save(output_path)
            self.logger.info(f"✅ Excel отчет успешно создан: {output_path}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"❌ Ошибка создания Excel отчета: {e}")
            raise
    
    def _add_headers(self, ws) -> None:
        """Добавляет строку заголовков с правильным форматированием."""
        headers = self.layout.get_column_headers()
        
        # Заголовки начинаются с позиции (start_row, start_col)
        for col_idx, header in enumerate(headers):
            cell = ws.cell(
                row=self.start_row, 
                column=self.start_col + col_idx, 
                value=header
            )
            
            # 6. Применяем новый цвет заголовков #FCE4D6 (Orange, Accent 2, Lighter 80%)
            cell.font = Font(bold=True, color="000000")  # Жирный черный текст
            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Новый оранжевый фон
            cell.alignment = Alignment(horizontal="center", vertical="center")  # По центру
            
            # Жирные границы для заголовков
            thick_border = Side(border_style="thick", color="000000")
            thin_border = Side(border_style="thin", color="000000")
            
            cell.border = Border(
                top=thick_border,
                left=thick_border if col_idx == 0 else thin_border,
                right=thick_border if col_idx == len(headers) - 1 else thin_border,
                bottom=thick_border  # Нижняя граница заголовков жирная
            )
    
    def _add_data_rows(self, ws, data: List[Dict[str, Any]]) -> None:
        """Добавляет строки данных с правильным форматированием."""
        
        for row_idx, record in enumerate(data):
            ws_row = self.start_row + 1 + row_idx  # +1 чтобы не перезаписать заголовки
            
            # Получаем данные строки в правильном порядке
            row_data = [
                record.get('account_number', ''),
                record.get('inn', ''),
                record.get('counterparty', ''),
                record.get('amount', ''),
                record.get('vat_amount', ''),
                record.get('invoice_date', ''),
                record.get('shipping_date', ''),
                record.get('payment_date', '')
            ]
            
            # Определяем цвет строки
            fill_color = self._get_row_color(record)
            
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(
                    row=ws_row,
                    column=self.start_col + col_idx,
                    value=value
                )
                
                # 1. ИСПРАВЛЕНИЕ: НЕ преобразуем номер счета в число, оставляем полный формат "XXXXX-XX"
                # Убираем преобразование номера счета, чтобы сохранить полный формат с дефисом
                
                # Цветовая заливка строки
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                # Специальное выравнивание для краткого отчета
                if col_idx == 4:  # Столбец "НДС" в кратком отчете
                    if str(value).lower() == "нет":
                        # Для "нет" - центрирование
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        # Для числовых значений - правое выравнивание
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx == 7:  # Столбец "Дата оплаты" в кратком отчете
                    # Для дат оплаты - центрирование
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Выравнивание по типу столбца для остальных
                    cell.alignment = self._get_column_alignment(col_idx)
                
                # Числовое форматирование
                cell.number_format = self._get_column_number_format(col_idx)
                
                # Обычные границы для данных
                thin_border = Side(border_style="thin", color="000000")
                thick_border = Side(border_style="thick", color="000000")
                
                cell.border = Border(
                    top=thin_border,
                    left=thick_border if col_idx == 0 else thin_border,
                    right=thick_border if col_idx == len(row_data) - 1 else thin_border,
                    bottom=thin_border
                )
    
    def _apply_data_table_borders(self, ws, data_rows: int) -> None:
        """4. Применяет жирную рамку только вокруг таблицы с данными (БЕЗ итогов)."""
        
        thick_border = Side(border_style="thick", color="000000")
        
        # Рамка заканчивается ПОСЛЕ последней строки данных
        last_data_row = self.start_row + data_rows  # заголовки + данные
        last_col = self.start_col + 7  # 8 столбцов = индекс 7
        
        # Жирная граница только вокруг таблицы с данными
        for row in range(self.start_row, last_data_row + 1):
            for col in range(self.start_col, last_col + 1):
                cell = ws.cell(row=row, column=col)
                
                border_left = thick_border if col == self.start_col else cell.border.left
                border_right = thick_border if col == last_col else cell.border.right  
                border_top = thick_border if row == self.start_row else cell.border.top
                border_bottom = thick_border if row == last_data_row else cell.border.bottom
                
                cell.border = Border(
                    left=border_left,
                    right=border_right,
                    top=border_top,
                    bottom=border_bottom
                )
    
    def _add_summary_section_new_format(self, ws, data: List[Dict[str, Any]]) -> None:
        """5. Добавляет итоги в новом формате как на скриншоте 04.png."""
        
        if not data:
            return
        
        # 🔧 ИСПРАВЛЕНИЕ: Используем правильные поля из process_invoice_record (amount, vat_amount)
        total_amount = sum(self._parse_amount(record.get('amount', 0)) for record in data)
        total_vat = sum(self._parse_amount(record.get('vat_amount', 0)) for record in data)
        
        # Вычисляем счета с НДС и без НДС (по vat_amount)
        no_vat_records = [r for r in data if self._parse_amount(r.get('vat_amount', 0)) == 0]
        with_vat_records = [r for r in data if self._parse_amount(r.get('vat_amount', 0)) > 0]
        
        no_vat_amount = sum(self._parse_amount(record.get('amount', 0)) for record in no_vat_records)
        with_vat_amount = sum(self._parse_amount(record.get('amount', 0)) for record in with_vat_records)
        
        # 2. Позиция для итогов (строка после данных + 1 пустая строка вместо 2)
        summary_start_row = self.start_row + len(data) + 2
        
        # 5. Новый формат итогов как на скриншоте 04.png
        summaries = [
            ("Всего счетов на сумму:", total_amount),
            ("Счетов без НДС на сумму:", no_vat_amount), 
            ("Счетов с НДС на сумму:", with_vat_amount),
            ("Всего НДС в счетах:", total_vat)
        ]
        
        for idx, (label, amount) in enumerate(summaries):
            current_row = summary_start_row + idx
            
            # 2. Подпись в столбце D (Контрагент) - обычный шрифт, выравнивание по правому краю
            label_cell = ws.cell(row=current_row, column=self.start_col + 2, value=label)
            label_cell.font = Font(bold=False)  # Убираем жирность
            label_cell.alignment = Alignment(horizontal="right")  # Правое выравнивание
            
            # 2. Сумма в столбце E (Сумма)
            amount_cell = ws.cell(row=current_row, column=self.start_col + 3, value=amount)
            amount_cell.alignment = Alignment(horizontal="right")
            amount_cell.number_format = '#,##0.00'
            
            # 2. Цвет и стиль значений: красный только для НДС, остальные черные жирные
            if "НДС в счетах" in label:
                amount_cell.font = Font(bold=True, color="FF0000")  # Красный и жирный только для НДС
            else:
                amount_cell.font = Font(bold=True, color="000000")  # Черный и жирный для остальных
        
        self.logger.info(f"📊 Новые итоги: {len(data)} счетов, всего: {total_amount:,.2f}, без НДС: {no_vat_amount:,.2f}, с НДС: {with_vat_amount:,.2f}, НДС: {total_vat:,.2f}")
    
    def _get_row_color(self, record: Dict[str, Any]) -> Optional[str]:
        """Определяет цвет строки по данным записи."""
        
        # Красный для неоплаченных
        if record.get('is_unpaid', False):
            return "FFC0CB"  # Красный для неоплаченных
            
        # Серый для записей без НДС
        if record.get('is_no_vat', False):
            return "D3D3D3"  # Серый для "Без НДС"
            
        return None  # Белый фон для обычных записей
    
    def _get_column_alignment(self, col_idx: int) -> Alignment:
        """Возвращает выравнивание для столбца по индексу."""
        
        # Столбцы для центрирования: Номер(0), ИНН(1), Дата счёта(5), Дата отгрузки(6), Дата оплаты(7)
        center_columns = [0, 1, 5, 6, 7]
        # Столбцы для правого выравнивания: Сумма(3), НДС(4)  
        right_columns = [3, 4]
         
        if col_idx in center_columns:
            return Alignment(horizontal="center", vertical="center")
        elif col_idx in right_columns:
            return Alignment(horizontal="right", vertical="center")
        else:
            return Alignment(horizontal="left", vertical="center")  # Контрагент слева
    
    def _get_column_number_format(self, col_idx: int) -> str:
        """Возвращает числовое форматирование для столбца."""
        
        # 1. ИСПРАВЛЕНИЕ: Номер остается как текст для сохранения дефиса
        # Убираем числовое форматирование для столбца "Номер", оставляем General
        
        # ИНН как число (столбец 1)
        if col_idx == 1:
            return '0'  # Целое число без разделителей
            
        # Суммы как числа (столбцы 3, 4)
        elif col_idx in [3, 4]:
            return '#,##0.00'  # Число с разделителями тысяч и 2 знака после запятой
            
        else:
            return 'General'  # Обычный формат для остальных (включая Номер)
    
    def _parse_amount(self, amount_str: str) -> float:
        """Парсит сумму из строки в число."""
        
        if not amount_str or amount_str == 'нет':
            return 0.0
            
        try:
            # Убираем все кроме цифр, точек и запятых
            clean_str = str(amount_str).replace(' ', '').replace(',', '.')
            return float(clean_str)
        except (ValueError, TypeError):
            return 0.0
    
    def _freeze_headers(self, ws) -> None:
        """Заморозка заголовков при прокручивании."""
        
        # Заморозка только строки заголовков, без заморозки столбцов
        # Устанавливаем заморозку на A3 - это заморозит только строки выше (заголовки), 
        # но позволит всем столбцам прокручиваться свободно
        freeze_cell = f"A{self.start_row + 1}"  # A3 - заморозка только строк, столбцы свободны
        ws.freeze_panes = freeze_cell
        self.logger.info(f"🧊 Заголовки заморожены на позиции: {freeze_cell} (только строки, столбцы свободны)")
    
    def _adjust_column_widths_auto(self, ws, data: List[Dict[str, Any]]) -> None:
        """1. Настройка ширины столбцов с автоподбором для "Контрагент", "Дата счёта", "Дата оплаты"."""
        
        # 3. Столбец A (пустой) делаем очень узким - примерно как высота строки
        ws.column_dimensions[get_column_letter(1)].width = 3  # Очень узкий столбец A
        
        # Базовые ширины столбцов
        column_widths = {
            self.start_col + 0: 12,  # Номер
            self.start_col + 1: 15,  # ИНН
            self.start_col + 2: 25,  # Контрагент (базовая ширина)
            self.start_col + 3: 15,  # Сумма
            self.start_col + 4: 12,  # НДС
            self.start_col + 5: 14,  # Дата счёта (базовая ширина)
            self.start_col + 6: 14,  # Дата отгрузки
            self.start_col + 7: 14,  # Дата оплаты (базовая ширина)
        }
        
        # 1. Автоподбор ширины для "Контрагент", "Дата счёта", "Дата оплаты"
        if data:
            # Анализируем длину контрагентов
            max_counterparty_len = max(len(str(record.get('counterparty', ''))) for record in data) if data else 0
            if max_counterparty_len > 25:
                column_widths[self.start_col + 2] = min(max_counterparty_len + 2, 50)  # Максимум 50
            
            # Анализируем даты (обычно одинаковые по длине, но проверим на всякий случай)
            max_invoice_date_len = max(len(str(record.get('invoice_date', ''))) for record in data) if data else 0
            max_payment_date_len = max(len(str(record.get('payment_date', ''))) for record in data) if data else 0
            
            if max_invoice_date_len > 14:
                column_widths[self.start_col + 5] = min(max_invoice_date_len + 2, 20)
            if max_payment_date_len > 14:
                column_widths[self.start_col + 7] = min(max_payment_date_len + 2, 20)
        
        # Применяем ширины
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
            
        self.logger.info(f"📏 Настроены ширины столбцов: A={3}, Контрагент={column_widths[self.start_col + 2]}, Даты={column_widths[self.start_col + 5]}")

    def _ensure_xlsx_extension(self, output_path: str) -> str:
        """
        Обеспечивает корректное расширение .xlsx для выходного файла.
        
        Args:
            output_path: Исходный путь к файлу
            
        Returns:
            Путь с правильным расширением .xlsx
        """
        path = Path(output_path)
        
        # Если нет расширения или расширение не .xlsx
        if path.suffix.lower() != '.xlsx':
            # Заменяем расширение на .xlsx
            new_path = path.with_suffix('.xlsx')
            self.logger.info(f"📝 Изменено расширение файла: {output_path} → {new_path}")
            return str(new_path)
        
        return output_path

    # ============================================================================
    # НОВЫЕ МЕТОДЫ ДЛЯ ДЕТАЛЬНОГО ОТЧЕТА - ФАЗА 4: EXCEL ГЕНЕРАЦИЯ
    # ============================================================================
    
    def create_detailed_report_sheet(
        self, 
        ws: Worksheet, 
        detailed_data: List[Dict[str, Any]]
    ) -> None:
        """
        Создает лист "Полный" с детальными данными товаров.
        
        Интегрируется с DataProcessor из Фазы 3 для отображения 
        товаров по счетам с зебра-эффектом группировки.
        
        Args:
            ws: Рабочий лист для детального отчета
            detailed_data: Данные товаров из format_products_for_excel()
        """
        try:
            self.logger.info(f"📋 Создание детального листа: {len(detailed_data)} товаров")
            
            # Записываем детальные данные с зебра-эффектом
            self.detailed_builder.write_detailed_data(ws, detailed_data)
            
            # 🔧 ИСПРАВЛЕНИЕ: Убираем итоги с листа "Полный" согласно требованию пользователя
            # summary_stats = self._calculate_detailed_summary(detailed_data)
            # self.detailed_builder.add_detailed_summary(ws, len(detailed_data), summary_stats)
            
            self.logger.info(f"✅ Детальный лист создан: {len(detailed_data)} товаров (без итогов)")
            
        except Exception as e:
            self.logger.error(f"❌ Ошибка создания детального листа: {e}")
            raise

    def create_multi_sheet_report(
        self, 
        brief_data: List[Dict[str, Any]], 
        detailed_data: List[Dict[str, Any]], 
        output_path: str
    ) -> str:
        """
        Создает двухлистовой Excel отчет: "Краткий" + "Полный".
        
        🔧 ИСПРАВЛЕНИЕ: Лист "Краткий" создается ТОЧНО ТАКИМ ЖЕ как однолистовой отчет.
        Используется стандартное форматирование ExcelReportGenerator для совместимости.
        
        Args:
            brief_data: Данные для краткого отчета (счета)
            detailed_data: Данные для детального отчета (товары)
            output_path: Путь для сохранения файла
            
        Returns:
            Путь к созданному двухлистовому отчету
        """
        try:
            # Обеспечиваем правильное расширение
            output_path = self._ensure_xlsx_extension(output_path)
            
            self.logger.info(f"📊 Создание двухлистового отчета: {len(brief_data)} счетов, {len(detailed_data)} товаров")
            
            # Создаем пустую книгу (НЕ используем MultiSheetBuilder для избежания конфликтов)
            wb = Workbook()
            
            # === ЛИСТ "КРАТКИЙ" - ТОЧНО КАК В ОДНОЛИСТОВОМ ===
            brief_ws = wb.active
            brief_ws.title = "Краткий"
            
            # Используем ВСЕ стандартные методы из create_report() для полной совместимости
            self._add_headers(brief_ws)  # ✅ Оранжевые заголовки как в однолистовом
            self._add_data_rows(brief_ws, brief_data)
            self._apply_data_table_borders(brief_ws, len(brief_data))
            self._add_summary_section_new_format(brief_ws, brief_data)  # ✅ Итоги с красным НДС
            self._freeze_headers(brief_ws)
            self._adjust_column_widths_auto(brief_ws, brief_data)
            
            # === ЛИСТ "ПОЛНЫЙ" - ИСПОЛЬЗУЕМ ДЕТАЛЬНЫЙ BUILDER ===
            detailed_ws = self.detailed_builder.create_detailed_worksheet(wb, "Полный")
            
            # Создаем детальный лист с товарами
            self.create_detailed_report_sheet(detailed_ws, detailed_data)
            
            # Сохраняем файл
            wb.save(output_path)
            
            self.logger.info(f"✅ Двухлистовой Excel отчет создан: {output_path}")
            self.logger.info(f"🎨 Лист 'Краткий': оранжевые заголовки, итоги с красным НДС (как в однолистовом)")
            self.logger.info(f"🎨 Лист 'Полный': зеленые заголовки, зебра-группировка товаров")
            return output_path
            
        except Exception as e:
            self.logger.error(f"❌ Ошибка создания двухлистового отчета: {e}")
            raise

    def _calculate_detailed_summary(self, detailed_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Вычисляет итоговую статистику для детального отчета.
        
        Args:
            detailed_data: Данные товаров
            
        Returns:
            Словарь с итоговой статистикой
        """
        # Подсчитываем уникальные счета
        unique_invoices = set()
        total_amount = 0.0
        
        for item in detailed_data:
            invoice_id = item.get('invoice_id')
            if invoice_id:
                unique_invoices.add(invoice_id)
            
            # Парсим сумму товара (удаляем пробелы и заменяем запятые)
            total_str = item.get('total_amount', '0')
            try:
                # Удаляем пробелы и заменяем запятую на точку
                clean_total = str(total_str).replace(' ', '').replace(',', '.')
                total_amount += float(clean_total)
            except (ValueError, TypeError):
                pass
        
        return {
            'total_invoices': len(unique_invoices),
            'total_products': len(detailed_data),
            'total_amount': total_amount
        }

    def generate_comprehensive_report(
        self, 
        brief_data: List[Dict[str, Any]], 
        product_data: Any,  # DetailedInvoiceData или processed data из DataProcessor
        output_path: str
    ) -> str:
        """
        Генерирует комплексный отчет с интеграцией всех фаз проекта.
        
        Объединяет:
        - API данные из Фазы 2 (Bitrix24Client)
        - Обработанные данные из Фазы 3 (DataProcessor)
        - Excel генерацию из Фазы 4 (ExcelReportGenerator)
        
        Args:
            brief_data: Краткие данные счетов
            product_data: Детальные данные товаров (из DataProcessor)
            output_path: Путь для сохранения
            
        Returns:
            Путь к созданному комплексному отчету
        """
        try:
            self.logger.info("🎯 Генерация комплексного отчета с интеграцией всех фаз")
            
            # Если product_data это результат из DataProcessor
            if hasattr(product_data, 'format_products_for_excel'):
                # Получаем данные в формате для Excel
                detailed_data = product_data.format_products_for_excel()
            elif isinstance(product_data, dict) and 'products' in product_data:
                # Если это grouped_data из group_products_by_invoice
                detailed_data = []
                for invoice_id, invoice_data in product_data.items():
                    for product in invoice_data.products:
                        if product.is_valid:
                            detailed_data.append({
                                'invoice_number': invoice_data.account_number,
                                'company_name': invoice_data.company_name or 'Не найдено',
                                'inn': invoice_data.inn or 'Не найдено',
                                'product_name': product.product_name,
                                'quantity': product.formatted_quantity,
                                'unit_measure': product.unit_measure,
                                'price': product.formatted_price,
                                'total_amount': product.formatted_total,
                                'invoice_id': invoice_id
                            })
            else:
                # Если это уже готовые данные для Excel
                detailed_data = product_data
            
            # Создаем двухлистовой отчет
            return self.create_multi_sheet_report(brief_data, detailed_data, output_path)
            
        except Exception as e:
            self.logger.error(f"❌ Ошибка генерации комплексного отчета: {e}")
            raise


class ReportGenerationError(Exception):
    """Custom exception for report generation errors."""
    pass


class ExcelReportBuilder:
    """
    High-level builder interface for Excel reports.
    
    Provides a convenient API for creating reports with different configurations.
    """
    
    def __init__(self):
        self.generator = ExcelReportGenerator()
    
    def build_invoice_report(self, invoices: List[Dict[str, Any]], 
                           output_path: str,
                           report_title: str = "Отчёт по счетам") -> str:
        """
        Build an invoice report.
        
        Args:
            invoices: List of invoice data
            output_path: Where to save the report
            report_title: Title for the report sheet
            
        Returns:
            Path to the generated report
        """
        try:
            self.generator.create_report(
                data=invoices,
                output_path=output_path
            )
            return output_path
        except Exception as e:
            raise ReportGenerationError(f"Failed to generate invoice report: {e}")
    
    def build_summary_report(self, invoices: List[Dict[str, Any]], 
                           output_path: str) -> Dict[str, Any]:
        """
        Build a summary report with statistics only.
        
        Args:
            invoices: List of invoice data
            output_path: Where to save the report
            
        Returns:
            Dictionary with summary statistics
        """
        try:
            # Calculate basic summary statistics
            summary = {
                'record_count': len(invoices),
                'total_without_vat': sum(record.get('amount', 0) for record in invoices),
                'total_with_vat': sum(record.get('amount', 0) + record.get('vat_amount', 0) for record in invoices)
            }
            
            # Generate Excel file
            self.generator.create_report(
                data=invoices,
                output_path=output_path
            )
            
            return summary
            
        except Exception as e:
            raise ReportGenerationError(f"Failed to generate summary report: {e}")
    
    def validate_data_for_report(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Validate data before report generation.
        
        Args:
            data: List of data to validate
            
        Returns:
            Dictionary with validation results
        """
        results = {
            'is_valid': True,
            'total_records': len(data),
            'valid_records': 0,
            'invalid_records': 0,
            'errors': []
        }
        
        for i, record in enumerate(data):
            try:
                # Basic validation - check if record has required fields
                required_fields = ['counterparty', 'amount']
                missing_fields = [field for field in required_fields if not record.get(field)]
                
                if missing_fields:
                    raise ValueError(f"Missing required fields: {missing_fields}")
                    
                results['valid_records'] += 1
                
            except Exception as e:
                results['invalid_records'] += 1
                results['errors'].append(f"Record {i + 1}: {str(e)}")
                results['is_valid'] = False
        
        return results
    
    # ============================================================================
    # НОВЫЕ МЕТОДЫ BUILDER ДЛЯ ДЕТАЛЬНОГО ОТЧЕТА - ФАЗА 4
    # ============================================================================
    
    def build_detailed_report(
        self, 
        brief_data: List[Dict[str, Any]], 
        detailed_data: List[Dict[str, Any]], 
        output_path: str
    ) -> str:
        """
        Создает детальный двухлистовой отчет.
        
        Args:
            brief_data: Данные для краткого отчета
            detailed_data: Данные для детального отчета
            output_path: Путь сохранения
            
        Returns:
            Путь к созданному отчету
        """
        try:
            return self.generator.create_multi_sheet_report(
                brief_data=brief_data,
                detailed_data=detailed_data,
                output_path=output_path
            )
        except Exception as e:
            raise ReportGenerationError(f"Failed to generate detailed report: {e}")
    
    def build_comprehensive_report(
        self, 
        brief_data: List[Dict[str, Any]], 
        product_data: Any, 
        output_path: str
    ) -> str:
        """
        Создает комплексный отчет с интеграцией всех фаз.
        
        Args:
            brief_data: Краткие данные счетов
            product_data: Данные товаров (из DataProcessor)
            output_path: Путь сохранения
            
        Returns:
            Путь к созданному отчету
        """
        try:
            return self.generator.generate_comprehensive_report(
                brief_data=brief_data,
                product_data=product_data,
                output_path=output_path
            )
        except Exception as e:
            raise ReportGenerationError(f"Failed to generate comprehensive report: {e}") 