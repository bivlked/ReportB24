"""
Unit tests for Excel layout module.

Tests report structure, column definitions, and worksheet layout
that matches the screenshot requirements.
"""

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.excel_generator.layout import (
    ColumnDefinition,
    ReportLayout,
    SummaryLayout,
    WorksheetBuilder
)


class TestColumnDefinition:
    """Test column definition dataclass."""
    
    def test_column_definition_creation(self):
        """Test ColumnDefinition dataclass creation."""
        col_def = ColumnDefinition(
            header="Test Header",
            width=15.5,
            alignment="center",
            data_key="test_key"
        )
        
        assert col_def.header == "Test Header"
        assert col_def.width == 15.5
        assert col_def.alignment == "center"
        assert col_def.data_key == "test_key"
    
    def test_column_definition_types(self):
        """Test ColumnDefinition field types."""
        col_def = ColumnDefinition(
            header="Header",
            width=10.0,
            alignment="left",
            data_key="key"
        )
        
        assert isinstance(col_def.header, str)
        assert isinstance(col_def.width, float)
        assert isinstance(col_def.alignment, str)
        assert isinstance(col_def.data_key, str)


class TestReportLayout:
    """Test report layout configuration."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.layout = ReportLayout()
    
    def test_layout_initialization(self):
        """Test ReportLayout initialization."""
        assert hasattr(self.layout, 'COLUMNS')
        assert hasattr(self.layout, 'HEADER_ROW')
        assert hasattr(self.layout, 'DATA_START_ROW')
        assert hasattr(self.layout, 'START_COLUMN')
        assert hasattr(self.layout, 'total_columns')
        
        assert self.layout.total_columns == len(self.layout.COLUMNS)
    
    def test_column_definitions_structure(self):
        """Test column definitions match screenshot requirements."""
        expected_columns = [
            ("Номер", 15.0, "center", "invoice_number"),
            ("ИНН", 15.0, "center", "inn"),
            ("Контрагент", 30.0, "left", "contractor_name"),
            ("Сумма", 18.0, "right", "total_amount"),
            ("НДС", 18.0, "center", "vat_amount"),
            ("Дата счёта", 15.0, "right", "invoice_date"),
            ("Дата отгрузки", 15.0, "right", "shipment_date"),
            ("Дата оплаты", 15.0, "right", "payment_date"),
        ]
        
        assert len(self.layout.COLUMNS) == 8
        
        for i, (header, width, alignment, data_key) in enumerate(expected_columns):
            col_def = self.layout.COLUMNS[i]
            assert isinstance(col_def, ColumnDefinition)
            assert col_def.header == header
            assert col_def.width == width
            assert col_def.alignment == alignment
            assert col_def.data_key == data_key
    
    def test_layout_constants(self):
        """Test layout positioning constants."""
        assert self.layout.HEADER_ROW == 2
        assert self.layout.DATA_START_ROW == 3
        assert self.layout.START_COLUMN == 2
    
    def test_setup_worksheet(self):
        """Test worksheet setup with column widths and row heights."""
        wb = Workbook()
        ws = wb.active
        
        self.layout.setup_worksheet(ws)
        
        # Test column widths
        for i, col_def in enumerate(self.layout.COLUMNS, start=self.layout.START_COLUMN):
            col_letter = ws.cell(row=1, column=i).column_letter
            assert ws.column_dimensions[col_letter].width == col_def.width
        
        # Test row height
        assert ws.row_dimensions[self.layout.HEADER_ROW].height == 20
        
        # Test freeze panes
        expected_freeze_cell = f'B3'  # START_COLUMN=2, DATA_START_ROW=3
        assert ws.freeze_panes == expected_freeze_cell
    
    def test_write_headers(self):
        """Test writing column headers to worksheet."""
        wb = Workbook()
        ws = wb.active
        
        self.layout.write_headers(ws)
        
        # Test header values
        for i, col_def in enumerate(self.layout.COLUMNS, start=self.layout.START_COLUMN):
            cell = ws.cell(row=self.layout.HEADER_ROW, column=i)
            assert cell.value == col_def.header
    
    def test_get_data_cell_position(self):
        """Test data cell position calculation."""
        # Test first data row, first column
        row, col = self.layout.get_data_cell_position(0, 0)
        assert row == self.layout.DATA_START_ROW  # 3
        assert col == self.layout.START_COLUMN    # 2
        
        # Test second data row, third column
        row, col = self.layout.get_data_cell_position(1, 2)
        assert row == self.layout.DATA_START_ROW + 1  # 4
        assert col == self.layout.START_COLUMN + 2    # 4
        
        # Test larger indices
        row, col = self.layout.get_data_cell_position(10, 5)
        assert row == self.layout.DATA_START_ROW + 10  # 13
        assert col == self.layout.START_COLUMN + 5     # 7
    
    def test_get_column_key(self):
        """Test getting data key for columns."""
        # Test valid columns
        assert self.layout.get_column_key(0) == "invoice_number"
        assert self.layout.get_column_key(1) == "inn"
        assert self.layout.get_column_key(2) == "contractor_name"
        assert self.layout.get_column_key(7) == "payment_date"
        
        # Test invalid columns
        assert self.layout.get_column_key(-1) == ""
        assert self.layout.get_column_key(8) == ""
        assert self.layout.get_column_key(100) == ""
    
    def test_get_column_alignment(self):
        """Test getting alignment for columns."""
        # Test valid columns
        assert self.layout.get_column_alignment(0) == "center"  # Номер
        assert self.layout.get_column_alignment(1) == "center"  # ИНН
        assert self.layout.get_column_alignment(2) == "left"    # Контрагент
        assert self.layout.get_column_alignment(3) == "right"   # Сумма
        assert self.layout.get_column_alignment(7) == "right"   # Дата оплаты
        
        # Test invalid columns
        assert self.layout.get_column_alignment(-1) == "left"
        assert self.layout.get_column_alignment(8) == "left"
        assert self.layout.get_column_alignment(100) == "left"


class TestSummaryLayout:
    """Test summary layout for totals section."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.layout = ReportLayout()
        self.summary = SummaryLayout(self.layout)
    
    def test_summary_initialization(self):
        """Test SummaryLayout initialization."""
        assert self.summary.layout == self.layout
    
    def test_add_summary_section(self):
        """Test adding summary section to worksheet."""
        wb = Workbook()
        ws = wb.active
        
        # Test data
        data_row_count = 5
        totals = {
            'amount_without_vat': '1 234 567,89',
            'amount_with_vat': '1 481 481,47'
        }
        
        self.summary.add_summary_section(ws, data_row_count, totals)
        
        # Calculate expected summary start row
        expected_start_row = self.layout.DATA_START_ROW + data_row_count + 2  # 3 + 5 + 2 = 10
        
        # Test record count
        count_cell = ws.cell(row=expected_start_row, column=self.layout.START_COLUMN)
        count_value_cell = ws.cell(row=expected_start_row, column=self.layout.START_COLUMN + 1)
        assert count_cell.value == "Всего записей:"
        assert count_value_cell.value == data_row_count
        
        # Test total without VAT
        without_vat_row = expected_start_row + 1
        without_vat_cell = ws.cell(row=without_vat_row, column=self.layout.START_COLUMN)
        without_vat_value_cell = ws.cell(row=without_vat_row, column=self.layout.START_COLUMN + 4)
        assert without_vat_cell.value == "Общая сумма без НДС:"
        assert without_vat_value_cell.value == totals['amount_without_vat']
        
        # Test total with VAT
        with_vat_row = expected_start_row + 2
        with_vat_cell = ws.cell(row=with_vat_row, column=self.layout.START_COLUMN)
        with_vat_value_cell = ws.cell(row=with_vat_row, column=self.layout.START_COLUMN + 6)
        assert with_vat_cell.value == "Общая сумма с НДС:"
        assert with_vat_value_cell.value == totals['amount_with_vat']
    
    def test_add_summary_section_partial_totals(self):
        """Test adding summary section with partial totals."""
        wb = Workbook()
        ws = wb.active
        
        # Test with only one total
        data_row_count = 3
        totals = {'amount_without_vat': '500 000,00'}
        
        self.summary.add_summary_section(ws, data_row_count, totals)
        
        expected_start_row = self.layout.DATA_START_ROW + data_row_count + 2
        
        # Test record count is always added
        count_cell = ws.cell(row=expected_start_row, column=self.layout.START_COLUMN)
        assert count_cell.value == "Всего записей:"
        
        # Test without VAT total is added
        without_vat_row = expected_start_row + 1
        without_vat_cell = ws.cell(row=without_vat_row, column=self.layout.START_COLUMN)
        assert without_vat_cell.value == "Общая сумма без НДС:"
        
        # Test with VAT total is not added when key is missing in totals
        with_vat_row = expected_start_row + 2
        with_vat_cell = ws.cell(row=with_vat_row, column=self.layout.START_COLUMN)
        # Should be None since 'amount_with_vat' key is not in totals dict
        assert with_vat_cell.value is None
    
    def test_add_summary_section_empty_totals(self):
        """Test adding summary section with empty totals."""
        wb = Workbook()
        ws = wb.active
        
        data_row_count = 0
        totals = {}
        
        self.summary.add_summary_section(ws, data_row_count, totals)
        
        expected_start_row = self.layout.DATA_START_ROW + data_row_count + 2
        
        # Record count should still be added
        count_cell = ws.cell(row=expected_start_row, column=self.layout.START_COLUMN)
        count_value_cell = ws.cell(row=expected_start_row, column=self.layout.START_COLUMN + 1)
        assert count_cell.value == "Всего записей:"
        assert count_value_cell.value == 0


class TestWorksheetBuilder:
    """Test worksheet builder class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.builder = WorksheetBuilder()
    
    def test_builder_initialization(self):
        """Test WorksheetBuilder initialization."""
        assert isinstance(self.builder.layout, ReportLayout)
        assert isinstance(self.builder.summary, SummaryLayout)
    
    def test_create_worksheet_new_workbook(self):
        """Test creating worksheet in new workbook."""
        wb = Workbook()
        ws = self.builder.create_worksheet(wb, "Test Sheet")
        
        assert ws.title == "Test Sheet"
        assert ws == wb.active
        
        # Test that layout configuration was applied
        # Check if headers were written
        for i, col_def in enumerate(self.builder.layout.COLUMNS, start=self.builder.layout.START_COLUMN):
            cell = ws.cell(row=self.builder.layout.HEADER_ROW, column=i)
            assert cell.value == col_def.header
    
    def test_create_worksheet_default_name(self):
        """Test creating worksheet with default name."""
        wb = Workbook()
        ws = self.builder.create_worksheet(wb)
        
        assert ws.title == "Отчёт"
    
    def test_calculate_totals(self):
        """Test calculating totals from data."""
        test_data = [
            {'amount_without_vat': 100.0, 'amount_with_vat': 120.0},
            {'amount_without_vat': 200.0, 'amount_with_vat': 240.0},
            {'amount_without_vat': 50.5, 'amount_with_vat': 60.6},
        ]
        
        totals = self.builder.calculate_totals(test_data)
        
        assert totals['count'] == 3
        assert totals['amount_without_vat'] == 350.5
        assert totals['amount_with_vat'] == 420.6
    
    def test_calculate_totals_mixed_types(self):
        """Test calculating totals with mixed data types."""
        test_data = [
            {'amount_without_vat': 100, 'amount_with_vat': 120},  # int
            {'amount_without_vat': 200.5, 'amount_with_vat': 240.6},  # float
            {'amount_without_vat': 'invalid', 'amount_with_vat': 'invalid'},  # string
            {},  # empty dict
        ]
        
        totals = self.builder.calculate_totals(test_data)
        
        assert totals['count'] == 4
        assert totals['amount_without_vat'] == 300.5  # Only numeric values summed
        assert totals['amount_with_vat'] == 360.6
    
    def test_calculate_totals_empty_data(self):
        """Test calculating totals with empty data."""
        totals = self.builder.calculate_totals([])
        
        assert totals['count'] == 0
        assert totals['amount_without_vat'] == 0.0
        assert totals['amount_with_vat'] == 0.0
    
    def test_calculate_totals_missing_fields(self):
        """Test calculating totals with missing amount fields."""
        test_data = [
            {'contractor_name': 'Test'},  # No amount fields
            {'amount_without_vat': 100},  # Missing with_vat
            {'amount_with_vat': 120},     # Missing without_vat
        ]
        
        totals = self.builder.calculate_totals(test_data)
        
        assert totals['count'] == 3
        assert totals['amount_without_vat'] == 100.0
        assert totals['amount_with_vat'] == 120.0


class TestLayoutIntegration:
    """Integration tests for layout module."""
    
    def test_complete_layout_workflow(self):
        """Test complete layout workflow from builder to summary."""
        builder = WorksheetBuilder()
        wb = Workbook()
        ws = builder.create_worksheet(wb, "Integration Test")
        
        # Test data
        test_data = [
            {'amount_without_vat': 1000.0, 'amount_with_vat': 1200.0},
            {'amount_without_vat': 2000.0, 'amount_with_vat': 2400.0},
        ]
        
        # Calculate totals
        totals = builder.calculate_totals(test_data)
        
        # Add summary section
        builder.summary.add_summary_section(ws, len(test_data), totals)
        
        # Verify complete structure
        # Headers should be present
        for i, col_def in enumerate(builder.layout.COLUMNS, start=builder.layout.START_COLUMN):
            cell = ws.cell(row=builder.layout.HEADER_ROW, column=i)
            assert cell.value == col_def.header
        
        # Summary should be present
        summary_start_row = builder.layout.DATA_START_ROW + len(test_data) + 2
        count_cell = ws.cell(row=summary_start_row, column=builder.layout.START_COLUMN)
        assert count_cell.value == "Всего записей:"
    
    def test_layout_coordinates_consistency(self):
        """Test that layout coordinates are consistent across components."""
        layout = ReportLayout()
        
        # Test that data start row is immediately after header row
        assert layout.DATA_START_ROW == layout.HEADER_ROW + 1
        
        # Test that column positions are consistent
        for col_index in range(len(layout.COLUMNS)):
            row, col = layout.get_data_cell_position(0, col_index)
            assert col == layout.START_COLUMN + col_index
            assert row == layout.DATA_START_ROW
    
    def test_screenshot_layout_requirements(self):
        """Test that layout matches screenshot requirements."""
        layout = ReportLayout()
        
        # Test Russian column headers match screenshots (обновлено под новую структуру)
        expected_headers = [
            "Номер", "ИНН", "Контрагент", "Сумма",
            "НДС", "Дата счёта", "Дата отгрузки", "Дата оплаты"
        ]
        
        actual_headers = [col.header for col in layout.COLUMNS]
        assert actual_headers == expected_headers
        
        # Test column widths are reasonable for content
        for col in layout.COLUMNS:
            assert col.width > 0
            assert col.width <= 50  # Reasonable maximum width
        
        # Test alignment matches screenshot requirements
        # Row numbers and INN should be centered
        assert layout.COLUMNS[0].alignment == "center"  # Номер
        assert layout.COLUMNS[1].alignment == "center"  # ИНН
        assert layout.COLUMNS[4].alignment == "center"  # НДС
        
        # Names should be left-aligned
        assert layout.COLUMNS[2].alignment == "left"    # Контрагент
        
        # Numbers and dates should be right-aligned
        assert layout.COLUMNS[3].alignment == "right"   # Сумма
        assert layout.COLUMNS[5].alignment == "right"   # Дата счёта
        assert layout.COLUMNS[6].alignment == "right"   # Дата отгрузки
        assert layout.COLUMNS[7].alignment == "right"   # Дата оплаты 