#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Тестовая версия правильного генератора отчетов
Фиксированные параметры, только 3 записи для быстрого тестирования
"""

import requests
import configparser
import time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def test_api_connection():
    """Тестирование подключения к API"""
    print("🔌 Тестирование подключения к Bitrix24 API...")
    
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')
    webhook_url = config.get('BitrixAPI', 'webhookurl')
    
    print(f"🔗 URL: {webhook_url}")
    
    # Тест 1: Получение Smart Invoices
    print("📋 Тест 1: Получение Smart Invoices...")
    params = {
        'entityTypeId': 31,
        'start': 0,
        'filter': {'!stageId': 'DT31_1:D'},
        'select': ['id', 'accountNumber', 'opportunity', 'taxValue']
    }
    
    try:
        resp = requests.post(f"{webhook_url}crm.item.list", json=params)
        if resp.status_code == 200:
            data = resp.json()
            items = data.get('result', {}).get('items', [])
            print(f"✅ Получено записей: {len(items)}")
            
            if items:
                # Берем первые 3 записи для тестирования
                test_invoices = items[:3]
                print("\n📊 Первые 3 записи:")
                for i, inv in enumerate(test_invoices, 1):
                    print(f"  {i}. ID: {inv.get('id')}, Номер: {inv.get('accountNumber')}")
                
                return test_invoices, webhook_url
            else:
                print("⚠️ Нет данных в ответе")
                return [], webhook_url
        else:
            print(f"❌ Ошибка API: {resp.status_code}")
            return [], webhook_url
            
    except Exception as e:
        print(f"💥 Ошибка соединения: {e}")
        return [], webhook_url


def test_requisites(invoice, webhook_url):
    """Тестирование получения реквизитов для одного счета"""
    acc_num = invoice.get('accountNumber', '')
    print(f"\n🔍 Тест реквизитов для счета: {acc_num}")
    
    try:
        # 1. Поиск счета по номеру
        print("  📋 Поиск счета...")
        resp = requests.get(f"{webhook_url}crm.item.list", params={
            'filter[accountNumber]': acc_num,
            'entityTypeId': 31
        })
        time.sleep(0.5)
        data = resp.json()
        
        items = data.get('result', {}).get('items', [])
        if not items:
            print("    ❌ Счет не найден")
            return "Не найден", "Нет данных"
        
        inv_id = items[0].get('id')
        print(f"    ✅ ID: {inv_id}")
        
        # 2. Поиск связи реквизитов
        print("  🔗 Поиск связей реквизитов...")
        link_resp = requests.get(f"{webhook_url}crm.requisite.link.list", params={
            'filter[ENTITY_TYPE_ID]': 31,
            'filter[ENTITY_ID]': inv_id
        })
        time.sleep(0.5)
        link_data = link_resp.json()
        link_items = link_data.get('result', [])
        
        if not link_items:
            print("    ❌ Нет связей реквизитов")
            return "Нет связей", "Нет связей"
        
        req_id = link_items[0].get('REQUISITE_ID')
        print(f"    ✅ Реквизит ID: {req_id}")
        
        # 3. Получение самого реквизита
        print("  📝 Получение реквизита...")
        req_resp = requests.post(f"{webhook_url}crm.requisite.get", json={"id": str(req_id)})
        time.sleep(0.5)
        req_data = req_resp.json()
        
        if not req_data.get('result'):
            print("    ❌ Ошибка получения реквизита")
            return "Ошибка реквизита", "Ошибка"
        
        fields = req_data['result']
        rq_inn = fields.get('RQ_INN', '')
        rq_company = fields.get('RQ_COMPANY_NAME', '')
        rq_name = fields.get('RQ_NAME', '')
        
        print(f"    📊 ИНН: {rq_inn}")
        print(f"    🏢 Компания: {rq_company}")
        print(f"    👤 Имя: {rq_name}")
        
        # Определение типа
        if rq_inn.isdigit():
            if len(rq_inn) == 10:
                result = (rq_company, rq_inn)
                print(f"    🏢 Тип: ООО/ЗАО")
            elif len(rq_inn) == 12:
                result = (f"ИП {rq_name}" if rq_name else "ИП (нет имени)", rq_inn)
                print(f"    👤 Тип: ИП")
            else:
                result = (rq_company, rq_inn)
                print(f"    ❓ Тип: Нестандартный ИНН")
        else:
            result = (rq_company, rq_inn)
            print(f"    ⚠️ Тип: ИНН не число")
        
        print(f"    ✅ Результат: '{result[0]}' | '{result[1]}'")
        return result
        
    except Exception as e:
        print(f"    💥 Ошибка: {e}")
        return "Ошибка", "Ошибка"


def create_test_report(invoices, webhook_url):
    """Создание тестового отчета"""
    print(f"\n📝 Создание тестового отчета ({len(invoices)} записей)")
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Тест отчета"
    
    # Заголовки как в ShortReport.py
    headers = ["Номер", "ИНН", "Контрагент", "Сумма", "НДС", "Дата счёта", "Дата отгрузки", "Дата оплаты"]
    ws.append(headers)
    
    # Цвета
    header_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
    red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Заголовок
    for cell in ws["1:1"]:
        cell.fill = header_fill
    
    row_index = 2
    for i, inv in enumerate(invoices, 1):
        print(f"\n📋 Обработка записи {i}/{len(invoices)}")
        
        acc_num = inv.get('accountNumber', '')
        opportunity = inv.get('opportunity', 0)
        tax_value = inv.get('taxValue', 0)
        
        # Форматирование суммы
        sum_val = f"{float(opportunity):,.2f}".replace(',', ' ').replace('.', ',')
        tax_text = f"{float(tax_value):,.2f}".replace(',', ' ').replace('.', ',') if float(tax_value) != 0 else "нет"
        
        # Получаем реквизиты
        comp_name, inn = test_requisites(inv, webhook_url)
        
        # Тестовые даты
        date_bill = "01.01.2022"
        ship_date = "02.01.2022"
        pay_date = "" if i == 1 else "03.01.2022"  # Первая запись без даты оплаты
        
        row_data = [acc_num, inn, comp_name, sum_val, tax_text, date_bill, ship_date, pay_date]
        ws.append(row_data)
        
        # Цветовая логика
        if pay_date == "":
            fill_color = red_fill
            color_desc = "🔴 Неоплаченный"
        elif tax_text == "нет":
            fill_color = grey_fill
            color_desc = "🔘 Без НДС"
        else:
            fill_color = white_fill
            color_desc = "⚪ Обычный"
        
        print(f"   🎨 Цвет: {color_desc}")
        
        # Применяем цвет к строке
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_index, column=col_idx)
            cell.fill = fill_color
        
        row_index += 1
    
    # Сохраняем
    Path('reports').mkdir(exist_ok=True)
    file_path = Path('reports') / 'Test_Correct_Report.xlsx'
    wb.save(file_path)
    
    print(f"\n✅ Тестовый отчет создан: {file_path}")
    return True


def main():
    """Главная функция"""
    print("=" * 60)
    print("🧪 ТЕСТ ПРАВИЛЬНОГО ГЕНЕРАТОРА ОТЧЕТОВ BITRIX24")
    print(f"⏰ Время: {datetime.now().strftime('%Y.%m.%d %H:%M:%S')}")
    print("=" * 60)
    
    try:
        # Тестируем подключение
        invoices, webhook_url = test_api_connection()
        
        if not invoices:
            print("❌ Нет данных для тестирования")
            return
        
        # Создаем тестовый отчет
        success = create_test_report(invoices, webhook_url)
        
        if success:
            print("\n🎉 ТЕСТ ЗАВЕРШЕН УСПЕШНО!")
            print("📁 Проверьте файл: reports/Test_Correct_Report.xlsx")
        else:
            print("\n❌ ОШИБКА ТЕСТИРОВАНИЯ")
            
    except Exception as e:
        print(f"\n💥 Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"\n⏰ Время завершения: {datetime.now().strftime('%Y.%m.%d %H:%M:%S')}")
    input("⏸️ Нажмите Enter для выхода...")


if __name__ == "__main__":
    main() 