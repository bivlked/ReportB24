#!/usr/bin/env python3
"""
Прямой генератор отчётов Bitrix24.
Простой подход без сложной архитектуры.
"""

import requests
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
import configparser
import os
from datetime import datetime
import time


def main():
    print("🚀 Прямой генератор отчётов Bitrix24")
    print("=" * 50)
    
    try:
        # Читаем config.ini
        print("📁 Чтение конфигурации...")
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')
        
        webhook_url = config.get('BitrixAPI', 'webhookurl').rstrip('/')
        save_folder = config.get('AppSettings', 'defaultsavefolder')
        filename = config.get('AppSettings', 'defaultfilename')
        
        print(f"⚙️  Webhook: {webhook_url[:50]}...")
        print(f"📂 Папка: {save_folder}")
        print(f"📄 Файл: {filename}")
        
        # Создаем папку
        os.makedirs(save_folder, exist_ok=True)
        save_path = os.path.join(save_folder, filename)
        
        # Простой тест API
        print("\n🧪 Тест API...")
        response = requests.get(f"{webhook_url}/profile", timeout=10)
        if response.status_code != 200:
            print(f"❌ API не работает: {response.status_code}")
            return False
        print("✅ API работает")
        
        # Получаем Smart Invoices
        print("\n📋 Получение Smart Invoices...")
        invoices_data = []
        start = 0
        
        while len(invoices_data) < 10:  # Ограничим первыми 10 для теста
            print(f"   Получение с позиции {start}...")
            
            response = requests.post(
                f"{webhook_url}/crm.item.list",
                json={
                    'entityTypeId': 31,
                    'start': start,
                    'limit': 10
                },
                timeout=30
            )
            
            if response.status_code != 200:
                print(f"❌ Ошибка получения данных: {response.status_code}")
                break
                
            data = response.json()
            items = data.get('result', {}).get('items', [])
            
            if not items:
                break
                
            invoices_data.extend(items)
            print(f"   Получено: {len(items)} записей")
            
            if len(items) < 10:  # Последняя страница
                break
                
            start += 10
            time.sleep(1)  # Rate limiting
        
        print(f"✅ Всего получено: {len(invoices_data)} Smart Invoices")
        
        if not invoices_data:
            print("❌ Данные не найдены")
            return False
        
        # Создаем Excel
        print(f"\n📊 Создание Excel отчёта...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Smart Invoices Report"
        
        # Заголовки
        headers = ["Номер", "ID", "Название", "Сумма", "НДС", "Дата создания", "Статус"]
        ws.append(headers)
        
        # Стили
        header_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
        center_align = Alignment(horizontal="center")
        
        # Форматируем заголовки
        for cell in ws[1]:
            cell.fill = header_fill
            cell.alignment = center_align
        
        # Добавляем данные
        for invoice in invoices_data:
            row = [
                invoice.get('accountNumber', ''),
                invoice.get('id', ''),
                invoice.get('title', ''),
                invoice.get('opportunity', 0),
                invoice.get('taxValue', 0),
                invoice.get('createdTime', ''),
                invoice.get('stageId', '')
            ]
            ws.append(row)
        
        # Автоподбор колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем
        wb.save(save_path)
        
        print(f"🎉 Отчёт создан!")
        print(f"📄 Путь: {save_path}")
        
        # Проверяем файл
        if os.path.exists(save_path):
            size = os.path.getsize(save_path)
            print(f"📦 Размер: {size:,} байт")
            print(f"📊 Записей: {len(invoices_data)}")
            return True
        else:
            print("❌ Файл не создался")
            return False
            
    except Exception as e:
        print(f"\n💥 Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = main()
    
    print("\n" + "=" * 50)
    if success:
        print("✅ ГОТОВО! Отчёт создан успешно!")
    else:
        print("❌ Ошибка создания отчёта")
    
    print("\n⏸️  Нажмите Enter для закрытия...")
    input() 