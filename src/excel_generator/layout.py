"""
Excel layout module for Bitrix24 report generation.

Defines the report structure, headers, and column configuration
that exactly matches the provided screenshots.
"""

from typing import List, Dict, Any, Tuple
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ColumnDefinition:
    """Definition for a single report column."""
    
    header: str           # Column header text
    width: float         # Column width in Excel units
    alignment: str       # Text alignment: 'left', 'center', 'right'
    data_key: str        # Key in data dict for this column


class ReportLayout:
    """
    Report layout configuration matching the screenshot requirements.
    
    Defines column structure, headers, and layout parameters
    for the Bitrix24 invoice report.
    """
    
    # Column definitions matching the screenshot
    COLUMNS = [
        ColumnDefinition(
            header="Номер",
            width=15.0,
            alignment="center",
            data_key="invoice_number"
        ),
        ColumnDefinition(
            header="ИНН",
            width=15.0,
            alignment="center",
            data_key="inn"
        ),
        ColumnDefinition(
            header="Контрагент",
            width=30.0,
            alignment="left",
            data_key="contractor_name"
        ),
        ColumnDefinition(
            header="Сумма",
            width=18.0,
            alignment="right",
            data_key="total_amount"
        ),
        ColumnDefinition(
            header="НДС",
            width=18.0,
            alignment="center",
            data_key="vat_amount"
        ),
        ColumnDefinition(
            header="Дата счёта",
            width=15.0,
            alignment="right",
            data_key="invoice_date"
        ),
        ColumnDefinition(
            header="Дата отгрузки",
            width=15.0,
            alignment="right",
            data_key="shipment_date"
        ),
        ColumnDefinition(
            header="Дата оплаты",
            width=15.0,
            alignment="right",
            data_key="payment_date"
        ),
    ]
    
    # Layout constants
    HEADER_ROW = 2          # Row for headers (1-based, skip empty row)
    DATA_START_ROW = 3      # First data row (1-based)
    START_COLUMN = 2        # Starting column (1-based, skip empty column)
    
    def __init__(self):
        self.total_columns = len(self.COLUMNS)
    
    def setup_worksheet(self, ws: Worksheet) -> None:
        """
        Set up worksheet with proper structure and column widths.
        
        Args:
            ws: OpenPyXL worksheet object
        """
        # Set column widths
        for i, col_def in enumerate(self.COLUMNS, start=self.START_COLUMN):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = col_def.width
        
        # Set row heights
        ws.row_dimensions[self.HEADER_ROW].height = 20  # Header row height
        
        # Freeze panes (freeze first row and first column)
        # This allows headers to stay visible when scrolling
        freeze_cell = ws.cell(row=self.DATA_START_ROW, column=self.START_COLUMN)
        ws.freeze_panes = freeze_cell
    
    def write_headers(self, ws: Worksheet) -> None:
        """
        Write column headers to the worksheet.
        
        Args:
            ws: OpenPyXL worksheet object
        """
        for i, col_def in enumerate(self.COLUMNS, start=self.START_COLUMN):
            cell = ws.cell(row=self.HEADER_ROW, column=i)
            cell.value = col_def.header
    
    def get_data_cell_position(self, row_index: int, column_index: int) -> Tuple[int, int]:
        """
        Get Excel cell position for data.
        
        Args:
            row_index: 0-based data row index
            column_index: 0-based column index
            
        Returns:
            Tuple of (row, column) in 1-based Excel coordinates
        """
        excel_row = self.DATA_START_ROW + row_index
        excel_col = self.START_COLUMN + column_index
        return excel_row, excel_col
    
    def get_column_key(self, column_index: int) -> str:
        """
        Get data key for a column.
        
        Args:
            column_index: 0-based column index
            
        Returns:
            Data key string for this column
        """
        if 0 <= column_index < len(self.COLUMNS):
            return self.COLUMNS[column_index].data_key
        return ""
    
    def get_column_alignment(self, column_index: int) -> str:
        """
        Get alignment for a column.
        
        Args:
            column_index: 0-based column index
            
        Returns:
            Alignment type: 'left', 'center', or 'right'
        """
        if 0 <= column_index < len(self.COLUMNS):
            return self.COLUMNS[column_index].alignment
        return "left"
    
    def get_column_headers(self) -> List[str]:
        """
        Get list of column headers.
        
        Returns:
            List of column header strings
        """
        return [col_def.header for col_def in self.COLUMNS]


class SummaryLayout:
    """
    Layout for summary/totals section at the bottom of the report.
    
    Provides structure for totals and statistics that appear
    below the main data table.
    """
    
    def __init__(self, layout: ReportLayout):
        self.layout = layout
    
    def add_summary_section(self, ws: Worksheet, data_row_count: int, 
                          totals: Dict[str, Any]) -> None:
        """
        Add summary section with totals below the data.
        
        Args:
            ws: OpenPyXL worksheet object
            data_row_count: Number of data rows written
            totals: Dictionary with total amounts
        """
        # Calculate summary start row (2 rows below last data)
        summary_start_row = self.layout.DATA_START_ROW + data_row_count + 2
        
        # Total count
        count_row = summary_start_row
        ws.cell(row=count_row, column=self.layout.START_COLUMN).value = "Всего записей:"
        ws.cell(row=count_row, column=self.layout.START_COLUMN + 1).value = data_row_count
        
        # Total amounts
        if 'amount_without_vat' in totals:
            amount_row = summary_start_row + 1
            ws.cell(row=amount_row, column=self.layout.START_COLUMN).value = "Общая сумма без НДС:"
            ws.cell(row=amount_row, column=self.layout.START_COLUMN + 4).value = totals['amount_without_vat']
        
        if 'amount_with_vat' in totals:
            total_row = summary_start_row + 2
            ws.cell(row=total_row, column=self.layout.START_COLUMN).value = "Общая сумма с НДС:"
            ws.cell(row=total_row, column=self.layout.START_COLUMN + 6).value = totals['amount_with_vat']


class WorksheetBuilder:
    """
    Builder class for creating and configuring Excel worksheets.
    
    Combines layout configuration with styling to create
    properly formatted worksheets.
    """
    
    def __init__(self):
        self.layout = ReportLayout()
        self.summary = SummaryLayout(self.layout)
    
    def create_worksheet(self, workbook: Workbook, sheet_name: str = "Отчёт") -> Worksheet:
        """
        Create and configure a new worksheet.
        
        Args:
            workbook: OpenPyXL workbook object
            sheet_name: Name for the worksheet
            
        Returns:
            Configured worksheet object
        """
        # Create or get worksheet
        if workbook.worksheets:
            ws = workbook.active
            ws.title = sheet_name
        else:
            ws = workbook.create_sheet(sheet_name)
        
        # Apply layout configuration
        self.layout.setup_worksheet(ws)
        self.layout.write_headers(ws)
        
        return ws
    
    def calculate_totals(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Calculate totals for the summary section.
        
        Args:
            data: List of data rows
            
        Returns:
            Dictionary with calculated totals
        """
        totals = {
            'count': len(data),
            'amount_without_vat': 0.0,
            'amount_with_vat': 0.0
        }
        
        for row in data:
            # Add amounts (handle both string and numeric values)
            if 'amount_without_vat' in row:
                amount = row['amount_without_vat']
                if isinstance(amount, (int, float)):
                    totals['amount_without_vat'] += amount
            
            if 'amount_with_vat' in row:
                amount = row['amount_with_vat']
                if isinstance(amount, (int, float)):
                    totals['amount_with_vat'] += amount
        
        return totals


# ============================================================================
# ДЕТАЛЬНЫЙ ОТЧЕТ LAYOUT - ФАЗА 4: EXCEL ГЕНЕРАЦИЯ
# ============================================================================

class DetailedReportLayout:
    """
    Layout configuration for detailed report with product breakdown.
    
    Создает макет для листа "Полный" с детальной информацией о товарах.
    Интегрируется с DataProcessor из Фазы 3 для обработки данных товаров.
    """
    
    # 🔧 ИСПРАВЛЕНИЕ: Обновленные определения колонок для детального отчета
    COLUMNS = [
        ColumnDefinition(
            header="Номер",  # Изменено с "Номер счёта"
            width=18.0,
            alignment="center",
            data_key="invoice_number"
        ),
        ColumnDefinition(
            header="ИНН",  # Поменяны местами с "Контрагент"
            width=15.0,
            alignment="center",
            data_key="inn"
        ),
        ColumnDefinition(
            header="Контрагент",  # Поменяны местами с "ИНН"
            width=25.0,
            alignment="left", 
            data_key="company_name"
        ),
        ColumnDefinition(
            header="Наименование товара",
            width=35.0,
            alignment="left",
            data_key="product_name"
        ),
        ColumnDefinition(
            header="Кол-во",
            width=12.0,
            alignment="center",  # Изменено с "right" на "center"
            data_key="quantity"
        ),
        ColumnDefinition(
            header="Цена",
            width=15.0,
            alignment="right",
            data_key="price"
        ),
        ColumnDefinition(
            header="Сумма",
            width=18.0,
            alignment="right",
            data_key="total_amount"
        ),
        ColumnDefinition(
            header="Сумма НДС",
            width=15.0,
            alignment="right",
            data_key="vat_amount"
        ),
    ]
    
    # Layout constants (same as brief report for consistency)
    HEADER_ROW = 2
    DATA_START_ROW = 3
    START_COLUMN = 2
    
    # Green header color from Hello World Excel Test
    HEADER_FILL_COLOR = "C6E0B4"  # Green color for headers
    
    def __init__(self):
        self.total_columns = len(self.COLUMNS)
    
    def setup_worksheet(self, ws: Worksheet) -> None:
        """
        Set up detailed worksheet with proper structure and styling.
        
        🔧 ИСПРАВЛЕНИЯ:
        - Столбец A узкий (как высота строки)
        - Убрана заморозка столбцов (только строка заголовков)
        - Автоширина для всех столбцов
        
        Args:
            ws: OpenPyXL worksheet object
        """
        from openpyxl.styles import PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        
        # 🔧 ИСПРАВЛЕНИЕ 1: Столбец A узкий (как высота строки)
        ws.column_dimensions['A'].width = 3  # Узкий столбец A
        
        # 🔧 ИСПРАВЛЕНИЕ 2: Автоширина для всех колонок данных
        # Временно устанавливаем базовые ширины, затем будет автоподбор
        for i, col_def in enumerate(self.COLUMNS, start=self.START_COLUMN):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = col_def.width
        
        # 🔧 ИСПРАВЛЕНИЕ: Унифицируем высоту строки заголовков между листами
        ws.row_dimensions[self.HEADER_ROW].height = 18  # Одинаковая высота с листом "Краткий"
        
        # 🔧 ИСПРАВЛЕНИЕ 3: Заморозка ТОЛЬКО строки заголовков (без столбцов)
        # Заморозка на A3 означает что заморожены строки 1-2, но столбцы свободны
        freeze_cell = f"A{self.DATA_START_ROW}"
        ws.freeze_panes = freeze_cell
        
        # Apply header styling (green background)
        header_fill = PatternFill(start_color=self.HEADER_FILL_COLOR, 
                                end_color=self.HEADER_FILL_COLOR, 
                                fill_type="solid")
        
        # Border style for professional look
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply styling to header row
        for i in range(self.START_COLUMN, self.START_COLUMN + self.total_columns):
            header_cell = ws.cell(row=self.HEADER_ROW, column=i)
            header_cell.fill = header_fill
            header_cell.border = border_style
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 🔧 УНИФИКАЦИЯ: Применяем тот же стиль шрифта как в кратком отчете
            from openpyxl.styles import Font
            header_cell.font = Font(bold=True, color="000000")  # Жирный черный текст
    
    def write_headers(self, ws: Worksheet) -> None:
        """
        Write column headers for detailed report.
        
        Args:
            ws: OpenPyXL worksheet object
        """
        for i, col_def in enumerate(self.COLUMNS, start=self.START_COLUMN):
            cell = ws.cell(row=self.HEADER_ROW, column=i)
            cell.value = col_def.header
    
    def apply_zebra_effect(self, ws: Worksheet, data_rows: List[Dict[str, Any]]) -> None:
        """
        Применение зебра-эффекта для группировки по счетам.
        
        Чередует цвет строк для каждого нового счета, чтобы
        визуально разделить товары разных счетов.
        
        Args:
            ws: OpenPyXL worksheet object  
            data_rows: Список строк данных с метаданными группировки
        """
        from openpyxl.styles import PatternFill
        
        # Colors for zebra effect (light gray for alternating invoices)
        zebra_color = "F2F2F2"  # Light gray
        zebra_fill = PatternFill(start_color=zebra_color, 
                               end_color=zebra_color, 
                               fill_type="solid")
        
        current_invoice_id = None
        use_zebra = False
        
        for row_idx, row_data in enumerate(data_rows):
            invoice_id = row_data.get('invoice_id')
            
            # Check if we're starting a new invoice group
            if invoice_id != current_invoice_id:
                current_invoice_id = invoice_id
                use_zebra = not use_zebra  # Toggle zebra effect
            
            # Apply zebra fill to this row if needed
            if use_zebra:
                excel_row = self.DATA_START_ROW + row_idx
                for col_idx in range(self.START_COLUMN, self.START_COLUMN + self.total_columns):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.fill = zebra_fill

    def apply_invoice_separator_borders(self, ws: Worksheet, data_rows: List[Dict[str, Any]]) -> None:
        """
        Применение толстых нижних границ для разделения счетов.
        
        Согласно Creative Phase решению: "Thick Bottom Border для последней строки каждого счета"
        обеспечивает максимальную визуальную ясность границ между счетами.
        
        Args:
            ws: OpenPyXL worksheet object
            data_rows: Список строк данных с метаданными группировки
        """
        from openpyxl.styles import Border, Side
        
        if not data_rows:
            return
        
        # Создаем стиль для границы между счетами
        thin_side = Side(border_style="thin", color="000000")
        medium_side = Side(border_style="medium", color="000000")
        
        separator_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=medium_side  # Толстая нижняя граница
        )
        
        current_invoice_id = None
        last_invoice_row = None
        
        # Находим последнюю строку каждого счета
        for row_idx, row_data in enumerate(data_rows):
            invoice_id = row_data.get('invoice_id')
            
            # Если начался новый счет и у нас есть предыдущий - применяем границу
            if invoice_id != current_invoice_id and last_invoice_row is not None:
                self._apply_separator_border_to_row(ws, last_invoice_row, separator_border)
            
            # Обновляем отслеживание
            if invoice_id != current_invoice_id:
                current_invoice_id = invoice_id
            
            last_invoice_row = row_idx
        
        # Применяем границу к последней строке последнего счета
        if last_invoice_row is not None:
            self._apply_separator_border_to_row(ws, last_invoice_row, separator_border)
    
    def _apply_separator_border_to_row(self, ws: 'Worksheet', row_idx: int, border) -> None:
        """
        Применяет границу разделения к конкретной строке.
        
        Args:
            ws: OpenPyXL worksheet object
            row_idx: Индекс строки данных (0-based)
            border: Стиль границы для применения
        """
        excel_row = self.DATA_START_ROW + row_idx
        
        # Применяем границу ко всем ячейкам строки
        for col_idx in range(len(self.COLUMNS)):
            excel_col = self.START_COLUMN + col_idx
            cell = ws.cell(row=excel_row, column=excel_col)
            
            # Просто применяем границу, не трогая заливку
            # (заливка уже установлена ранее через zebra-стилизацию)
            cell.border = border

    def get_data_cell_position(self, row_index: int, column_index: int) -> Tuple[int, int]:
        """
        Get Excel cell position for detailed data.
        
        Args:
            row_index: 0-based data row index
            column_index: 0-based column index
            
        Returns:
            Tuple of (row, column) in 1-based Excel coordinates
        """
        excel_row = self.DATA_START_ROW + row_index
        excel_col = self.START_COLUMN + column_index
        return excel_row, excel_col
    
    def get_column_key(self, column_index: int) -> str:
        """Get data key for a column in detailed report."""
        if 0 <= column_index < len(self.COLUMNS):
            return self.COLUMNS[column_index].data_key
        return ""
    
    def get_column_alignment(self, column_index: int) -> str:
        """Get alignment for a column in detailed report."""
        if 0 <= column_index < len(self.COLUMNS):
            return self.COLUMNS[column_index].alignment
        return "left"
    
    def get_column_headers(self) -> List[str]:
        """Get list of column headers for detailed report."""
        return [col_def.header for col_def in self.COLUMNS]


class DetailedWorksheetBuilder:
    """
    Builder for creating detailed report worksheets.
    
    Создает и настраивает листы для детального отчета с товарами.
    Интегрируется с DataProcessor для обработки данных.
    """
    
    def __init__(self):
        self.layout = DetailedReportLayout()
    
    def create_detailed_worksheet(
        self, 
        workbook: Workbook, 
        sheet_name: str = "Полный"
    ) -> Worksheet:
        """
        Create and configure detailed worksheet.
        
        Args:
            workbook: OpenPyXL workbook object
            sheet_name: Name for the detailed worksheet
            
        Returns:
            Configured detailed worksheet object
        """
        # Create new worksheet (don't replace existing ones)
        ws = workbook.create_sheet(sheet_name)
        
        # Apply detailed layout configuration
        self.layout.setup_worksheet(ws)
        self.layout.write_headers(ws)
        
        return ws
    
    def write_detailed_data(
        self, 
        ws: Worksheet, 
        data_rows: List[Dict[str, Any]]
    ) -> None:
        """
        Write detailed product data to worksheet.
        
        Записывает данные товаров с применением зебра-эффекта
        для группировки по счетам.
        
        Args:
            ws: OpenPyXL worksheet object
            data_rows: List of formatted product data from DataProcessor
        """
        from openpyxl.styles import Alignment, Border, Side
        
        # Border for data cells
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write data rows
        for row_idx, row_data in enumerate(data_rows):
            for col_idx, col_def in enumerate(self.layout.COLUMNS):
                excel_row, excel_col = self.layout.get_data_cell_position(row_idx, col_idx)
                
                cell = ws.cell(row=excel_row, column=excel_col)
                cell.value = row_data.get(col_def.data_key, "")
                
                # Apply alignment
                cell.alignment = Alignment(horizontal=col_def.alignment, vertical="center")
                
                # 🔧 УНИФИКАЦИЯ: Применяем числовое форматирование как в кратком отчете
                cell.number_format = self._get_detailed_column_number_format(col_idx)
                
                # Apply border
                cell.border = border_style
        
        # Apply zebra effect after writing all data
        self.layout.apply_zebra_effect(ws, data_rows)
        
        # Apply thick borders between invoices
        self.layout.apply_invoice_separator_borders(ws, data_rows)
        
        # 🔧 УНИФИКАЦИЯ: Применяем жирные границы вокруг таблицы как в кратком отчете
        self._apply_detailed_table_borders(ws, len(data_rows))
        
        # 🔧 ИСПРАВЛЕНИЕ: Автоподбор ширины столбцов после записи данных
        self._adjust_detailed_column_widths(ws, data_rows)
    
    def _get_detailed_column_number_format(self, col_idx: int) -> str:
        """
        🔧 УНИФИКАЦИЯ: Числовое форматирование для детального отчета
        
        Применяет такое же форматирование как в кратком отчете:
        - ИНН (индекс 2): '0' - целое число
        - Кол-во, Цена, Сумма (индексы 4, 6, 7): '#,##0.00' - числа с разделителями
        - Остальные: 'General' - обычный формат
        
        Args:
            col_idx: Индекс столбца (0-based)
            
        Returns:
            Строка формата числа для Excel
        """
        # 🔧 ИСПРАВЛЕНИЕ: Обновленное соответствие столбцов детального отчета:
        # 0: Номер (текст)
        # 1: ИНН (число) - поменялись местами с Контрагент
        # 2: Контрагент (текст) - поменялись местами с ИНН
        # 3: Наименование товара (текст)
        # 4: Кол-во (ЦЕЛОЕ число) - требование пользователя
        # 5: Ед. изм. (текст)
        # 6: Цена (число с 2 знаками)
        # 7: Сумма (число с 2 знаками)
        
        if col_idx == 1:  # ИНН (теперь второй столбец)
            return '0'  # Целое число без разделителей (как в кратком отчете)
        elif col_idx == 4:  # Кол-во - ЦЕЛЫЕ числа (без дробной части)
            return '0'  # Целое число без дробной части
        elif col_idx in [6, 7]:  # Цена, Сумма
            return '#,##0.00'  # Число с разделителями тысяч и 2 знака после запятой
        else:
            return 'General'  # Обычный формат для остальных
    
    def _apply_detailed_table_borders(self, ws: Worksheet, data_rows: int) -> None:
        """
        🔧 УНИФИКАЦИЯ: Применяет жирную рамку вокруг таблицы детального отчета
        
        Точно такая же логика как в _apply_data_table_borders() краткого отчета.
        
        Args:
            ws: Рабочий лист
            data_rows: Количество строк данных
        """
        from openpyxl.styles import Border, Side
        
        thick_border = Side(border_style="thick", color="000000")
        
        # Рамка заканчивается ПОСЛЕ последней строки данных
        last_data_row = self.layout.DATA_START_ROW + data_rows - 1  # заголовки + данные  
        last_col = self.layout.START_COLUMN + self.layout.total_columns - 1  # последний столбец
        
        # Жирная граница только вокруг таблицы с данными (БЕЗ итогов)
        for row in range(self.layout.HEADER_ROW, last_data_row + 1):
            for col in range(self.layout.START_COLUMN, last_col + 1):
                cell = ws.cell(row=row, column=col)
                
                border_left = thick_border if col == self.layout.START_COLUMN else cell.border.left
                border_right = thick_border if col == last_col else cell.border.right  
                border_top = thick_border if row == self.layout.HEADER_ROW else cell.border.top
                border_bottom = thick_border if row == last_data_row else cell.border.bottom
                
                cell.border = Border(
                    left=border_left,
                    right=border_right,
                    top=border_top,
                    bottom=border_bottom
                )
    
    def add_detailed_summary(
        self, 
        ws: Worksheet, 
        data_row_count: int,
        summary_stats: Dict[str, Any]
    ) -> None:
        """
        Add summary section for detailed report.
        
        Args:
            ws: OpenPyXL worksheet object
            data_row_count: Number of product rows written
            summary_stats: Summary statistics (total products, invoices, etc.)
        """
        # Calculate summary start row
        summary_start_row = self.layout.DATA_START_ROW + data_row_count + 2
        
        # Products count
        ws.cell(row=summary_start_row, column=self.layout.START_COLUMN).value = "Всего товаров:"
        ws.cell(row=summary_start_row, column=self.layout.START_COLUMN + 1).value = data_row_count
        
        # Invoices count
        if 'total_invoices' in summary_stats:
            invoices_row = summary_start_row + 1
            ws.cell(row=invoices_row, column=self.layout.START_COLUMN).value = "Всего счетов:"
            ws.cell(row=invoices_row, column=self.layout.START_COLUMN + 1).value = summary_stats['total_invoices']
    
    def _adjust_detailed_column_widths(self, ws: Worksheet, data_rows: List[Dict[str, Any]]) -> None:
        """
        🔧 ИСПРАВЛЕНИЕ: Автоподбор ширины столбцов для детального отчета
        
        Анализирует фактические данные и устанавливает оптимальную ширину
        для всех столбцов (как требовал пользователь).
        
        Args:
            ws: Рабочий лист
            data_rows: Список строк данных для анализа
        """
        from openpyxl.utils import get_column_letter
        
        if not data_rows:
            return
        
        # Анализируем длину данных в каждом столбце
        max_lengths = {}
        
        for col_idx, col_def in enumerate(self.layout.COLUMNS):
            data_key = col_def.data_key
            header_length = len(col_def.header)
            
            # Находим максимальную длину для этого столбца
            max_data_length = 0
            for row in data_rows:
                value = str(row.get(data_key, ""))
                max_data_length = max(max_data_length, len(value))
            
            # Учитываем и заголовок, и данные + небольшой отступ
            optimal_width = max(header_length, max_data_length) + 2
            
            # Применяем ограничения по типу столбца
            if data_key == "inn":
                optimal_width = max(optimal_width, 12)  # Минимум для ИНН
                optimal_width = min(optimal_width, 20)  # Максимум для ИНН
            elif data_key == "company_name":
                optimal_width = max(optimal_width, 20)  # Минимум для контрагента
                optimal_width = min(optimal_width, 50)  # Максимум для контрагента
            elif data_key == "product_name":
                optimal_width = max(optimal_width, 25)  # Минимум для товара
                optimal_width = min(optimal_width, 60)  # Максимум для товара
            elif data_key in ["price", "total_amount", "vat_amount"]:
                optimal_width = max(optimal_width, 15)  # Минимум для денежных полей
                optimal_width = min(optimal_width, 25)  # Максимум для денежных полей
            else:
                optimal_width = min(optimal_width, 30)  # Общий максимум
            
            max_lengths[col_idx] = optimal_width
        
        # Применяем вычисленные ширины
        for col_idx, optimal_width in max_lengths.items():
            excel_col = self.layout.START_COLUMN + col_idx
            col_letter = get_column_letter(excel_col)
            ws.column_dimensions[col_letter].width = optimal_width
        
        # Логирование результатов
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"📏 Автоподбор ширины столбцов детального отчета:")
        for col_idx, optimal_width in max_lengths.items():
            col_name = self.layout.COLUMNS[col_idx].header
            logger.info(f"   {col_name}: {optimal_width}")


class MultiSheetBuilder:
    """
    Builder for creating multi-sheet reports (Brief + Detailed).
    
    Создает Excel файл с двумя листами: "Краткий" и "Полный".
    Координирует работу обычного и детального макетов.
    """
    
    def __init__(self):
        self.brief_builder = WorksheetBuilder()
        self.detailed_builder = DetailedWorksheetBuilder()
    
    def create_multi_sheet_workbook(self) -> Workbook:
        """
        Create workbook with both brief and detailed sheets.
        
        Returns:
            Workbook with "Краткий" and "Полный" sheets
        """
        workbook = Workbook()
        
        # Remove default sheet if exists
        if workbook.worksheets:
            workbook.remove(workbook.active)
        
        # Create brief sheet first
        brief_ws = self.brief_builder.create_worksheet(workbook, "Краткий")
        
        # Create detailed sheet
        detailed_ws = self.detailed_builder.create_detailed_worksheet(workbook, "Полный")
        
        return workbook
    
    def get_brief_worksheet(self, workbook: Workbook) -> Worksheet:
        """Get brief worksheet from multi-sheet workbook."""
        return workbook["Краткий"]
    
    def get_detailed_worksheet(self, workbook: Workbook) -> Worksheet:
        """Get detailed worksheet from multi-sheet workbook."""
        return workbook["Полный"] 