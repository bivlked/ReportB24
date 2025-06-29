#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Оптимизированный генератор отчетов Bitrix24 Excel
С подробным логированием и ограничением для тестирования
"""

import requests
import configparser
import time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class OptimizedBitrix24Reporter:
    """Оптимизированный генератор отчетов с логированием"""
    
    def __init__(self, config_path="config.ini", max_records=10):
        """Инициализация с ограничением записей для тестирования"""
        print(f"⏰ Время инициализации: {datetime.now().strftime('%Y.%m.%d %H:%M:%S')}")
        self.max_records = max_records
        self.config = configparser.ConfigParser()
        self.config.read(config_path, encoding='utf-8')
        
        # Получаем настройки из config.ini
        self.webhook_url = self.config.get('BitrixAPI', 'webhookurl')
        self.save_folder = self.config.get('AppSettings', 'defaultsavefolder', fallback='reports')
        self.file_name = self.config.get('AppSettings', 'defaultfilename', fallback='Optimized_Report.xlsx')
        
        # Период из конфига
        start_date_str = self.config.get('ReportPeriod', 'startdate', fallback='01.01.2022')
        end_date_str = self.config.get('ReportPeriod', 'enddate', fallback='31.12.2022')
        
        self.start_date = datetime.strptime(start_date_str, "%d.%m.%Y").date()
        self.end_date = datetime.strptime(end_date_str, "%d.%m.%Y").date()
        
        # Создаем папку для отчетов
        Path(self.save_folder).mkdir(exist_ok=True)
        
        print(f"🔗 Webhook URL: {self.webhook_url}")
        print(f"📅 Период: {start_date_str} - {end_date_str}")
        print(f"📁 Папка сохранения: {self.save_folder}")
        print(f"📄 Имя файла: {self.file_name}")
        print(f"🔢 Ограничение записей: {self.max_records} (для тестирования)")

    def get_company_info_by_invoice(self, invoice_number):
        """Получение реквизитов контрагента с детальным логированием"""
        print(f"  🔍 Получение реквизитов для счета #{invoice_number}...")
        
        try:
            # 1. Ищем счёт по номеру
            print(f"    📋 Поиск счета...")
            resp = requests.get(f"{self.webhook_url}crm.item.list", params={
                'filter[accountNumber]': invoice_number,
                'entityTypeId': 31
            })
            time.sleep(0.5)
            data = resp.json()

            items = data.get('result', {}).get('items', [])
            if not items:
                print(f"    ❌ Счёт не найден")
                return "Счет не найден", "Нет данных"

            inv_id = items[0].get('id')
            print(f"    ✅ Счет найден, ID: {inv_id}")

            # 2. Ищем привязку реквизитов
            print(f"    🔗 Поиск связи реквизитов...")
            link_resp = requests.get(f"{self.webhook_url}crm.requisite.link.list", params={
                'filter[ENTITY_TYPE_ID]': 31,
                'filter[ENTITY_ID]': inv_id
            })
            time.sleep(0.5)
            link_data = link_resp.json()
            link_items = link_data.get('result', [])
            
            if not link_items:
                print(f"    ⚠️ Нет связанных реквизитов")
                return "Нет реквизитов", "Нет реквизитов"

            req_id = link_items[0].get('REQUISITE_ID')
            print(f"    ✅ Найден реквизит ID: {req_id}")

            # 3. Получаем сам реквизит
            print(f"    📝 Получение данных реквизита...")
            req_resp = requests.post(f"{self.webhook_url}crm.requisite.get", json={"id": str(req_id)})
            time.sleep(0.5)
            req_data = req_resp.json()
            
            if not req_data.get('result'):
                print(f"    ❌ Ошибка получения реквизита")
                return "Ошибка реквизита", "Ошибка реквизита"

            fields = req_data['result']
            rq_inn = fields.get('RQ_INN', '')
            rq_company = fields.get('RQ_COMPANY_NAME', '')
            rq_name = fields.get('RQ_NAME', '')
            
            print(f"    📊 ИНН: {rq_inn}, Компания: {rq_company}")

            # 4. Логика определения типа по ИНН
            if rq_inn.isdigit():
                if len(rq_inn) == 10:
                    result = (rq_company, rq_inn)  # ООО/ЗАО
                    print(f"    🏢 Определен как ООО/ЗАО")
                elif len(rq_inn) == 12:
                    result = (f"ИП {rq_name}" if rq_name else "ИП (нет имени)", rq_inn)  # ИП
                    print(f"    👤 Определен как ИП")
                else:
                    result = (rq_company, rq_inn)
                    print(f"    ❓ Нестандартный ИНН")
            else:
                result = (rq_company, rq_inn)
                print(f"    ⚠️ ИНН не является числом")
            
            print(f"    ✅ Результат: {result[0]} | {result[1]}")
            return result

        except Exception as e:
            print(f"    💥 Ошибка: {e}")
            return "Ошибка", "Ошибка"

    def get_smart_invoices_limited(self):
        """Получение ограниченного количества Smart Invoices для тестирования"""
        print(f"🔍 Получение Smart Invoices (ограничение: {self.max_records} записей)")
        url = f"{self.webhook_url}crm.item.list"
        
        params = {
            'entityTypeId': 31,
            'start': 0,
            'filter': {
                '!stageId': 'DT31_1:D'  # исключаем удаленные
            },
            'select': [
                'id', 'accountNumber', 'statusId', 'dateBill', 'price',
                'UFCRM_SMART_INVOICE_1651168135187',  # Дата отгрузки
                'UFCRM_626D6ABE98692',               # Дата оплаты
                'begindate', 'opportunity', 'stageId', 'taxValue'
            ]
        }
        
        print(f"📡 Отправка API запроса...")
        resp = requests.post(url, json=params)
        time.sleep(0.5)
        
        if resp.status_code != 200:
            print(f"❌ Ошибка API: {resp.status_code} - {resp.text}")
            return []

        data = resp.json()
        all_items = data.get('result', {}).get('items', [])
        print(f"📊 Получено всего записей: {len(all_items)}")
        
        # Ограничиваем количество для тестирования
        limited_items = all_items[:self.max_records]
        print(f"🔢 Используем для обработки: {len(limited_items)} записей")

        # Фильтрация по дате отгрузки
        filtered = []
        for inv in limited_items:
            ship_date_str = inv.get('UFCRM_SMART_INVOICE_1651168135187')
            if ship_date_str:
                try:
                    d = datetime.fromisoformat(ship_date_str.replace('Z', '+00:00')).date()
                    if self.start_date <= d <= self.end_date:
                        filtered.append(inv)
                except ValueError:
                    pass

        print(f"🎯 Отфильтровано по дате отгрузки: {len(filtered)} записей")
        return filtered

    def format_date(self, date_str):
        """Форматирование даты"""
        if not date_str:
            return ""
        try:
            d = datetime.fromisoformat(date_str.replace('Z', '+00:00')).date()
            return d.strftime("%d.%m.%Y")
        except:
            return ""

    def format_currency(self, amount):
        """Форматирование валюты"""
        try:
            return f"{float(amount):,.2f}".replace(',', ' ').replace('.', ',')
        except:
            return "0,00"

    def generate_report(self):
        """Генерация отчета с подробным логированием"""
        print("\n🚀 ЗАПУСК ГЕНЕРАЦИИ ОТЧЕТА")
        print("=" * 50)
        
        # Получаем данные
        invoices = self.get_smart_invoices_limited()
        if not invoices:
            print("❌ Нет данных для отчета")
            return False

        print(f"\n📝 ОБРАБОТКА ДАННЫХ ({len(invoices)} записей)")
        print("-" * 30)

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Счета (Smart Invoices)"

        # Заголовки
        headers = ["Номер", "ИНН", "Контрагент", "Сумма", "НДС", "Дата счёта", "Дата отгрузки", "Дата оплаты"]
        ws.append(headers)

        # Цвета
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        header_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")

        row_index = 2
        
        for i, inv in enumerate(invoices, 1):
            print(f"\n📋 Обработка записи {i}/{len(invoices)}")
            print(f"   ID: {inv.get('id')}, Номер: {inv.get('accountNumber')}")
                
            acc_num = inv.get('accountNumber', '')
            sum_val = self.format_currency(inv.get('opportunity', 0))
            tax_val = float(inv.get('taxValue', 0))
            tax_text = self.format_currency(tax_val) if tax_val != 0 else "нет"

            date_bill = self.format_date(inv.get('begindate'))
            ship_date = self.format_date(inv.get('UFCRM_SMART_INVOICE_1651168135187'))
            pay_date_str = inv.get('UFCRM_626D6ABE98692')
            pay_date = self.format_date(pay_date_str) if pay_date_str else ""

            # Получаем реквизиты
            comp_name, inn = self.get_company_info_by_invoice(acc_num)

            row_data = [acc_num, inn, comp_name, sum_val, tax_text, date_bill, ship_date, pay_date]
            ws.append(row_data)

            # Цветовая логика
            if pay_date == "":
                fill_color = red_fill
                color_desc = "🔴 Неоплаченный"
            elif tax_text == "нет":
                fill_color = grey_fill
                color_desc = "🔘 Без НДС"
            else:
                fill_color = white_fill
                color_desc = "⚪ Обычный"
            
            print(f"   🎨 Цвет строки: {color_desc}")

            # Применяем цвет
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_index, column=col_idx)
                cell.fill = fill_color

            row_index += 1

        print(f"\n🎨 ФОРМАТИРОВАНИЕ EXCEL")
        print("-" * 20)

        # Форматирование (упрощенное для тестирования)
        for cell in ws["1:1"]:
            cell.fill = header_fill

        # Сохраняем файл
        final_path = Path(self.save_folder) / self.file_name
        wb.save(final_path)
        
        print(f"\n✅ ОТЧЕТ СОЗДАН УСПЕШНО!")
        print(f"📁 Файл: {final_path}")
        print(f"📊 Записей обработано: {len(invoices)}")
        return True


def main():
    """Главная функция запуска"""
    print("=" * 60)
    print("🎯 ОПТИМИЗИРОВАННЫЙ ГЕНЕРАТОР ОТЧЕТОВ BITRIX24")
    print("=" * 60)
    
    try:
        # Спрашиваем количество записей для тестирования
        max_records = input("📝 Введите количество записей для обработки (по умолчанию 5): ").strip()
        if not max_records:
            max_records = 5
        else:
            max_records = int(max_records)
        
        reporter = OptimizedBitrix24Reporter(max_records=max_records)
        success = reporter.generate_report()
        
        if success:
            print("\n🎉 ОТЧЕТ УСПЕШНО СОЗДАН!")
        else:
            print("\n❌ ОШИБКА СОЗДАНИЯ ОТЧЕТА")
            
    except Exception as e:
        print(f"\n💥 Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
    
    input("\n⏸️ Нажмите Enter для выхода...")


if __name__ == "__main__":
    main() 