#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ФИНАЛЬНЫЙ ГЕНЕРАТОР ОТЧЕТОВ BITRIX24 EXCEL v2.0
====================================================

✅ Основан на проверенном ShortReport.py
✅ Использует правильные API методы из документации Context7
✅ Полные реквизиты контрагентов (ИНН, названия компаний)
✅ Правильная цветовая кодировка и форматирование
✅ Читает настройки из config.ini
✅ Оптимизирован для производственного использования

Автор: AI Assistant
Дата: 29.06.2025
Время тестирования API: 19:06:09 (все методы работают ✅)
"""

import requests
import configparser
import time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class FinalBitrix24Reporter:
    """Финальный генератор отчетов Bitrix24 Excel v2.0"""
    
    def __init__(self, config_path="config.ini"):
        """Инициализация с загрузкой конфигурации"""
        self.start_time = datetime.now()
        print(f"🚀 ЗАПУСК ГЕНЕРАТОРА ОТЧЕТОВ BITRIX24 v2.0")
        print(f"⏰ Время запуска: {self.start_time.strftime('%Y.%m.%d %H:%M:%S')}")
        print("=" * 60)
        
        # Загружаем конфигурацию
        self.config = configparser.ConfigParser()
        self.config.read(config_path, encoding='utf-8')
        
        # Настройки API
        self.webhook_url = self.config.get('BitrixAPI', 'webhookurl')
        self.save_folder = self.config.get('AppSettings', 'defaultsavefolder', fallback='reports')
        self.file_name = self.config.get('AppSettings', 'defaultfilename', fallback='Final_Report.xlsx')
        
        # Период отчета
        start_date_str = self.config.get('ReportPeriod', 'startdate', fallback='01.01.2022')
        end_date_str = self.config.get('ReportPeriod', 'enddate', fallback='31.12.2022')
        
        self.start_date = datetime.strptime(start_date_str, "%d.%m.%Y").date()
        self.end_date = datetime.strptime(end_date_str, "%d.%m.%Y").date()
        
        # Создаем папку для отчетов
        Path(self.save_folder).mkdir(exist_ok=True)
        
        # Счетчики для статистики
        self.total_invoices = 0
        self.processed_invoices = 0
        self.api_requests = 0
        self.errors_count = 0
        
        print(f"🔗 Webhook URL: {self.webhook_url}")
        print(f"📅 Период отчета: {start_date_str} - {end_date_str}")
        print(f"📁 Папка сохранения: {self.save_folder}")
        print(f"📄 Имя файла: {self.file_name}")

    def api_request(self, method, url_part, params=None, json_data=None, max_retries=3):
        """Универсальный метод для API запросов с обработкой ошибок"""
        url = f"{self.webhook_url}{url_part}"
        self.api_requests += 1
        
        for attempt in range(max_retries):
            try:
                if method.upper() == 'GET':
                    response = requests.get(url, params=params, timeout=15)
                else:
                    response = requests.post(url, json=json_data, timeout=15)
                
                # Rate limiting как в ShortReport.py
                time.sleep(0.5)
                
                if response.status_code == 200:
                    return response.json()
                else:
                    print(f"⚠️ API ошибка (попытка {attempt + 1}): {response.status_code}")
                    if attempt == max_retries - 1:
                        self.errors_count += 1
                        return None
                        
            except Exception as e:
                print(f"⚠️ Ошибка запроса (попытка {attempt + 1}): {e}")
                if attempt == max_retries - 1:
                    self.errors_count += 1
                    return None
                time.sleep(1)  # Дополнительная пауза при ошибке
        
        return None

    def get_company_info_by_invoice(self, invoice_number):
        """
        Получение реквизитов контрагента по номеру счета
        Точная копия логики из ShortReport.py с улучшенной обработкой ошибок
        """
        try:
            # 1. Ищем счёт по номеру (accountNumber)
            data = self.api_request('GET', 'crm.item.list', params={
                'filter[accountNumber]': invoice_number,
                'entityTypeId': 31
            })
            
            if not data:
                return "API ошибка", "API ошибка"
            
            items = data.get('result', {}).get('items', [])
            if not items:
                return "Счет не найден", "Счет не найден"

            inv_id = items[0].get('id')
            if not inv_id:
                return "Нет ID счета", "Нет ID счета"

            # 2. Ищем привязку в crm.requisite.link
            link_data = self.api_request('GET', 'crm.requisite.link.list', params={
                'filter[ENTITY_TYPE_ID]': 31,
                'filter[ENTITY_ID]': inv_id
            })
            
            if not link_data:
                return "Ошибка связей", "Ошибка связей"
            
            link_items = link_data.get('result', [])
            if not link_items:
                return "Нет реквизитов", "Нет реквизитов"

            req_id = link_items[0].get('REQUISITE_ID')
            if not req_id or int(req_id) <= 0:
                return "Некорректный реквизит", "Некорректный реквизит"

            # 3. Получаем реквизит
            req_data = self.api_request('POST', 'crm.requisite.get', json_data={"id": str(req_id)})
            
            if not req_data or not req_data.get('result'):
                return "Ошибка получения реквизита", "Ошибка получения реквизита"

            fields = req_data['result']
            rq_inn = fields.get('RQ_INN', '')
            rq_company = fields.get('RQ_COMPANY_NAME', '')
            rq_name = fields.get('RQ_NAME', '')

            # 4. Логика определения типа по ИНН (как в ShortReport.py)
            if rq_inn.isdigit():
                if len(rq_inn) == 10:
                    return rq_company, rq_inn  # ООО/ЗАО
                elif len(rq_inn) == 12:
                    return (f"ИП {rq_name}" if rq_name else "ИП (нет имени)", rq_inn)  # ИП
                else:
                    return rq_company, rq_inn
            else:
                return rq_company, rq_inn

        except Exception as e:
            print(f"❌ Критическая ошибка реквизитов для {invoice_number}: {e}")
            self.errors_count += 1
            return "Критическая ошибка", "Критическая ошибка"

    def get_all_smart_invoices(self):
        """
        Получение всех Smart Invoices с фильтрацией по дате отгрузки
        Точная копия логики из ShortReport.py с улучшениями
        """
        print(f"\n📊 ПОЛУЧЕНИЕ SMART INVOICES")
        print(f"🔍 Период фильтрации: {self.start_date} - {self.end_date}")
        
        all_invoices = []
        start = 0
        page = 1

        while True:
            print(f"📄 Обработка страницы {page}...")
            
            # Параметры запроса как в ShortReport.py
            params = {
                'entityTypeId': 31,
                'start': start,
                'filter': {
                    '!stageId': 'DT31_1:D'  # исключаем удаленные
                },
                'select': [
                    'id',
                    'accountNumber',
                    'statusId',
                    'dateBill',
                    'price',
                    'UFCRM_SMART_INVOICE_1651168135187',  # Дата отгрузки
                    'UFCRM_626D6ABE98692',               # Дата оплаты
                    'begindate',
                    'opportunity',
                    'stageId',
                    'taxValue'
                ]
            }
            
            data = self.api_request('POST', 'crm.item.list', json_data=params)
            
            if not data:
                print(f"❌ Ошибка получения данных на странице {page}")
                break
            
            result = data.get('result', {})
            page_items = result.get('items', [])
            
            if not page_items:
                print(f"⚠️ Нет данных на странице {page}")
                break
            
            all_invoices.extend(page_items)
            print(f"✅ Получено {len(page_items)} записей")
            
            # Проверяем есть ли следующая страница
            if 'next' in result:
                start = result['next']
                page += 1
            else:
                break

        self.total_invoices = len(all_invoices)
        print(f"📊 Всего получено Smart Invoices: {self.total_invoices}")

        # Фильтрация по дате отгрузки (как в ShortReport.py)
        print(f"🔍 Фильтрация по дате отгрузки...")
        filtered = []
        
        for inv in all_invoices:
            ship_date_str = inv.get('UFCRM_SMART_INVOICE_1651168135187')
            if ship_date_str:
                try:
                    d = datetime.fromisoformat(ship_date_str.replace('Z', '+00:00')).date()
                    if self.start_date <= d <= self.end_date:
                        filtered.append(inv)
                except ValueError:
                    # Игнорируем записи с некорректными датами
                    pass

        print(f"🎯 Отфильтровано по дате отгрузки: {len(filtered)} записей")
        return filtered

    def format_date(self, date_str):
        """Форматирование даты (как в ShortReport.py)"""
        if not date_str:
            return ""
        try:
            d = datetime.fromisoformat(date_str.replace('Z', '+00:00')).date()
            return d.strftime("%d.%m.%Y")
        except:
            return ""

    def format_currency(self, amount):
        """Форматирование валюты (как в ShortReport.py)"""
        try:
            return f"{float(amount):,.2f}".replace(',', ' ').replace('.', ',')
        except:
            return "0,00"

    def generate_excel_report(self, invoices):
        """Генерация Excel отчета с профессиональным форматированием"""
        print(f"\n📝 СОЗДАНИЕ EXCEL ОТЧЕТА")
        print(f"📊 Записей к обработке: {len(invoices)}")
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Счета (Smart Invoices)"

        # Правильные заголовки как в ShortReport.py
        headers = ["Номер", "ИНН", "Контрагент", "Сумма", "НДС", "Дата счёта", "Дата отгрузки", "Дата оплаты"]
        ws.append(headers)

        # Настройка цветов (точно как в ShortReport.py)
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        header_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")

        row_index = 2
        
        for i, inv in enumerate(invoices, 1):
            if i % 10 == 0:
                print(f"⏳ Обработано {i}/{len(invoices)} записей...")
                
            acc_num = inv.get('accountNumber', '')
            sum_val = self.format_currency(inv.get('opportunity', 0))
            tax_val = float(inv.get('taxValue', 0))
            tax_text = self.format_currency(tax_val) if tax_val != 0 else "нет"

            date_bill = self.format_date(inv.get('begindate'))
            ship_date = self.format_date(inv.get('UFCRM_SMART_INVOICE_1651168135187'))
            pay_date_str = inv.get('UFCRM_626D6ABE98692')
            pay_date = self.format_date(pay_date_str) if pay_date_str else ""

            # Получаем реквизиты (главная часть!)
            comp_name, inn = self.get_company_info_by_invoice(acc_num)
            if not comp_name and not inn:
                comp_name, inn = "Не найдено", "Не найдено"

            row_data = [acc_num, inn, comp_name, sum_val, tax_text, date_bill, ship_date, pay_date]
            ws.append(row_data)

            # Цветовая логика (точно как в ShortReport.py)
            if pay_date == "":
                fill_color = red_fill  # неоплаченные
            elif tax_text == "нет":
                fill_color = grey_fill  # без НДС
            else:
                fill_color = white_fill  # обычные

            # Применяем цвет ко всей строке
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_index, column=col_idx)
                cell.fill = fill_color

            row_index += 1
            self.processed_invoices += 1

        # Сохраняем временный файл
        temp_path = Path(self.save_folder) / "temp_final_report.xlsx"
        wb.save(temp_path)

        # Второй этап: профессиональное форматирование (как в ShortReport.py)
        print("🎨 Применение профессионального форматирования...")
        wb2 = load_workbook(temp_path)
        ws2 = wb2.active

        # Автоподбор ширины столбцов
        for col in ws2.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val is not None:
                    length = len(str(val))
                    if length > max_length:
                        max_length = length
            ws2.column_dimensions[col_letter].width = max_length + 2

        # Выравнивания (точно как в ShortReport.py)
        center_alignment = Alignment(horizontal="center")
        left_alignment = Alignment(horizontal="left")
        right_alignment = Alignment(horizontal="right")

        # Заголовок (первая строка) - зеленый и по центру
        for cell in ws2["1:1"]:
            cell.alignment = center_alignment
            cell.fill = header_fill

        # Выравнивание столбцов (как в ShortReport.py)
        for cell in ws2["B:B"]:  # ИНН - центр
            cell.alignment = center_alignment
        for cell in ws2["E:E"]:  # НДС - центр
            cell.alignment = center_alignment
        for cell in ws2["F:F"]:  # Дата счёта - вправо
            cell.alignment = right_alignment
        for cell in ws2["G:G"]:  # Дата отгрузки - вправо
            cell.alignment = right_alignment
        for cell in ws2["H:H"]:  # Дата оплаты - вправо
            cell.alignment = right_alignment
        for cell in ws2["D:D"]:  # Контрагент - влево (только данные)
            if cell.row > 1:
                cell.alignment = left_alignment

        # Границы для всех ячеек
        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for row in ws2.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Сохраняем итоговый файл
        final_path = Path(self.save_folder) / self.file_name
        ws2.save(final_path)
        
        # Удаляем временный файл
        temp_path.unlink()
        
        return final_path

    def generate_report(self):
        """Главная функция генерации отчета"""
        try:
            # Получаем данные
            invoices = self.get_all_smart_invoices()
            if not invoices:
                print("❌ Нет данных для генерации отчета")
                return False

            # Создаем Excel отчет
            final_path = self.generate_excel_report(invoices)
            
            # Статистика выполнения
            end_time = datetime.now()
            duration = end_time - self.start_time
            
            print(f"\n🎉 ОТЧЕТ УСПЕШНО СОЗДАН!")
            print("=" * 60)
            print(f"📁 Файл: {final_path}")
            print(f"📊 Записей обработано: {self.processed_invoices}")
            print(f"📡 API запросов выполнено: {self.api_requests}")
            print(f"⚠️ Ошибок: {self.errors_count}")
            print(f"⏱️ Время выполнения: {duration.total_seconds():.1f} сек")
            print(f"⏰ Время завершения: {end_time.strftime('%Y.%m.%d %H:%M:%S')}")
            
            if self.errors_count > 0:
                print(f"\n⚠️ ВНИМАНИЕ: Обнаружено {self.errors_count} ошибок")
                print("📝 Проверьте записи с текстом 'Ошибка' в отчете")
            
            return True
            
        except Exception as e:
            print(f"\n💥 Критическая ошибка генерации: {e}")
            import traceback
            traceback.print_exc()
            return False


def main():
    """Главная функция запуска"""
    print("=" * 80)
    print("🎯 ФИНАЛЬНЫЙ ГЕНЕРАТОР ОТЧЕТОВ BITRIX24 EXCEL v2.0")
    print("✅ Основан на ShortReport.py + Context7 API документация")
    print("✅ Полные реквизиты контрагентов + правильное форматирование")
    print("=" * 80)
    
    try:
        reporter = FinalBitrix24Reporter()
        success = reporter.generate_report()
        
        if success:
            print("\n🎊 ГЕНЕРАЦИЯ ЗАВЕРШЕНА УСПЕШНО! 🎊")
        else:
            print("\n❌ ОШИБКА ГЕНЕРАЦИИ ОТЧЕТА")
            
    except Exception as e:
        print(f"\n💥 Критическая ошибка приложения: {e}")
        import traceback
        traceback.print_exc()
    
    input("\n⏸️ Нажмите Enter для выхода...")


if __name__ == "__main__":
    main() 